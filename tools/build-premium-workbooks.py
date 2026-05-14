#!/usr/bin/env python3
"""Build premium restaurant growth and finance workbooks.

The generator intentionally keeps formulas inside Excel, not hidden in Python,
so each workbook remains auditable by operators, finance leads, and reviewers.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import zipfile
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
GROWTH_DIR = ROOT / "public" / "downloads" / "restaurant-growth"
FINANCE_DIR = ROOT / "public" / "downloads" / "restaurant-finance"
RUN_DATE = date.today()
OUTPUT_DIR = ROOT / "outputs" / f"{RUN_DATE:%Y%m%d}-premium-workbooks"
QA_JSON = OUTPUT_DIR / "qa-report.json"
QA_MD = OUTPUT_DIR / "qa-report.md"
SUMMARY_MD = OUTPUT_DIR / "summary-report.md"
TEMPLATE_DATE = RUN_DATE.strftime("%d-%b-%Y")
VERSION = "v2.0"

COLORS = {
    "navy": "0F172A",
    "charcoal": "111827",
    "canvas": "F8FAFC",
    "panel": "FFFFFF",
    "panel_alt": "EEF2F7",
    "input": "E7F1FF",
    "formula": "F8FAFC",
    "linked": "EAF7EF",
    "amber": "D4875A",
    "purple": "C9A2D4",
    "green": "10B981",
    "red": "EF4444",
    "yellow": "FEF3C7",
    "border": "CBD5E1",
    "border_soft": "E2E8F0",
    "muted": "64748B",
    "white": "FFFFFF",
}

FONTS = {
    "base": "Inter",
    "heading": "Inter Tight",
    "mono": "JetBrains Mono",
}

THIN = Side(style="thin", color=COLORS["border_soft"])
MEDIUM = Side(style="medium", color=COLORS["border"])
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor=COLORS["navy"])
SECTION_FILL = PatternFill("solid", fgColor=COLORS["panel_alt"])
INPUT_FILL = PatternFill("solid", fgColor=COLORS["input"])
FORMULA_FILL = PatternFill("solid", fgColor=COLORS["formula"])
LINK_FILL = PatternFill("solid", fgColor=COLORS["linked"])
YELLOW_FILL = PatternFill("solid", fgColor=COLORS["yellow"])

STATUS_VALUES = ["Not Started", "In Progress", "Ready", "Approved", "Live", "Complete", "At Risk", "Blocked"]
PRIORITY_VALUES = ["Critical", "High", "Medium", "Low"]
MONTH_VALUES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
OWNER_VALUES = [
    "Founder",
    "Marketing Lead",
    "Operations Lead",
    "Finance Lead",
    "Store Manager",
    "Franchisee",
    "Agency",
]
CHANNEL_VALUES = [
    "Meta",
    "Google",
    "TikTok",
    "Instagram",
    "Email",
    "SMS",
    "WhatsApp",
    "In-store",
    "Aggregator",
    "Influencer",
    "PR",
    "OOH",
]

SOURCE_LIBRARY = [
    (
        "Financial model standards",
        "FAST Standard",
        "https://fast-standard.org/",
        "Flexible, appropriate, structured, transparent model architecture.",
    ),
    (
        "Spreadsheet controls",
        "ICAEW financial modelling guidance",
        "https://www.icaew.com/technical/business/financial-management/financial-modelling-and-forecasting/getting-your-financial-models-right",
        "Documentation, checks, restrictions, and model review practices.",
    ),
    (
        "Restaurant dashboard KPIs",
        "Toast reporting dashboard",
        "https://support.toasttab.com/en/article/How-to-Use-the-Toast-Reporting-Dashboard",
        "Sales, labor, guest count, and menu performance as core restaurant reporting areas.",
    ),
    (
        "Marketing ROI",
        "WebstaurantStore restaurant marketing ROI",
        "https://www.webstaurantstore.com/article/384/restaurant-marketing-roi.html",
        "ROI should use net return and avoid vanity metrics.",
    ),
    (
        "Local SEO / GBP",
        "Google Business Profile local ranking",
        "https://support.google.com/business/answer/7091?hl=en-en",
        "Relevance, distance, prominence, complete info, reviews, photos.",
    ),
    (
        "GBP performance metrics",
        "Google Business Profile performance",
        "https://support.google.com/business/answer/9918094?hl=en-en",
        "Views, searches, calls, website clicks, and direction requests.",
    ),
    (
        "Menu engineering",
        "LibreTexts menu engineering",
        "https://workforce.libretexts.org/Bookshelves/Food_Production_Service_and_Culinary_Arts/Restaurant_Design%3A_Concept_to_Customer_%28Thibodeaux%29/12%3A_Restaurant_Analysis/12.01%3A_Menu_Engineering",
        "Contribution margin, food cost, RevPASH, Stars/Plowhorses/Puzzles/Dogs.",
    ),
    (
        "Delivery menu optimization",
        "DoorDash / Technomic delivery menu optimization",
        "https://assets.ctfassets.net/trvmqu12jq2l/2JVMSuZPRd1LYkHNHxHK8j/2473b685d3e1069d20adbfcfdefd253c/DoorDash_Technomic_Delivery_Menu_Optimization_2024_Report.pdf",
        "Photos, descriptions, customization, and delivery-specific menu quality signals.",
    ),
    (
        "Franchise audit governance",
        "ActionCard franchise audit features",
        "https://actioncardapp.com/",
        "Custom audits, action plans, visual evidence, dashboards, issue resolution.",
    ),
]


@dataclass
class SampleRow:
    record_id: str
    month: str
    dimension: str
    item: str
    owner: str
    channel: str
    plan: float
    actual: float
    value: float
    cost: float
    status: str
    priority: str
    evidence: str
    extras: dict[str, Any] = field(default_factory=dict)


@dataclass
class WorkbookSpec:
    filename: str
    folder: str
    title: str
    subtitle: str
    category: str
    audience: str
    management_question: str
    dimension_label: str
    item_label: str
    channel_label: str
    plan_label: str
    actual_label: str
    value_label: str
    cost_label: str
    efficiency_label: str
    chart_title: str
    gross_margin: float
    min_roi: float
    target_score: float
    assumptions: list[tuple[str, Any, str, str, str, str]]
    definitions: list[tuple[str, str, str]]
    actions: list[tuple[str, str, str, str]]
    extra_headers: list[str]
    rows: list[SampleRow]

    @property
    def output_dir(self) -> Path:
        return GROWTH_DIR if self.folder == "growth" else FINANCE_DIR

    @property
    def path(self) -> Path:
        return self.output_dir / self.filename

    @property
    def slug(self) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "_", self.filename.replace(".xlsx", "")).strip("_")


def money(value: float) -> float:
    return float(value)


def row(
    record_id: str,
    month: str,
    dimension: str,
    item: str,
    owner: str,
    channel: str,
    plan: float,
    actual: float,
    value: float,
    cost: float,
    status: str,
    priority: str,
    evidence: str,
    **extras: Any,
) -> SampleRow:
    return SampleRow(record_id, month, dimension, item, owner, channel, plan, actual, value, cost, status, priority, evidence, extras)


def make_specs() -> list[WorkbookSpec]:
    return [
        WorkbookSpec(
            filename="annual-marketing-budget-planner.xlsx",
            folder="growth",
            title="Annual Marketing Budget Control Model",
            subtitle="Planned vs actual marketing spend, revenue contribution, CAC, ROAS, utilization, and variance control.",
            category="Budget control",
            audience="Founders, marketing heads, finance leads, and multi-store operators.",
            management_question="Is marketing spend controlled, efficient, and tied to revenue contribution?",
            dimension_label="Channel / Budget Line",
            item_label="Campaign / Activity",
            channel_label="Primary Channel",
            plan_label="Planned Budget AED",
            actual_label="Actual Spend AED",
            value_label="Attributed Revenue AED",
            cost_label="Total Marketing Cost AED",
            efficiency_label="Contribution ROI",
            chart_title="Monthly planned vs actual spend",
            gross_margin=0.62,
            min_roi=0.25,
            target_score=0.85,
            assumptions=[
                ("Marketing budget as % of sales", 0.06, "%", "Management target", "Caps total spend", "Finance Lead"),
                ("Paid media CAC target", 38, "AED / new customer", "Internal benchmark", "Flags inefficient acquisition", "Marketing Lead"),
                ("Material overspend threshold", 0.10, "%", "FP&A policy", "Triggers variance commentary", "Finance Lead"),
                ("Minimum ROAS reference", 3.0, "x", "Operator assumption", "Channel decision guardrail", "Marketing Lead"),
            ],
            definitions=[
                ("Budget utilization", "Actual spend divided by planned budget.", "=Actual Spend / Planned Budget"),
                ("CAC", "Total marketing cost divided by new customers acquired.", "=Total Cost / New Customers"),
                ("ROAS", "Attributed revenue divided by marketing cost.", "=Attributed Revenue / Total Marketing Cost"),
                ("Contribution ROI", "Gross profit from attributed revenue less marketing cost, divided by marketing cost.", "=(Revenue x Gross Margin - Cost) / Cost"),
            ],
            actions=[
                ("Paid media overspend", "Pause low-margin ad sets and reallocate budget to strongest contribution ROI channel.", "Marketing Lead", "High"),
                ("CAC drift", "Add weekly CAC review by channel before approving next media flight.", "Finance Lead", "Medium"),
                ("Weak attribution", "Tag campaigns consistently and separate direct, aggregator, and CRM revenue.", "Marketing Lead", "High"),
            ],
            extra_headers=["New Customers", "CAC AED", "ROAS x", "Budget Owner Sign-off"],
            rows=[
                row("BUD-001", "Jan", "Paid Social", "New Year bundle push", "Marketing Lead", "Meta", 18000, 19400, 72000, 21400, "Complete", "High", "Ads export + POS tag", **{"New Customers": 510, "Budget Owner Sign-off": "Yes"}),
                row("BUD-002", "Feb", "Search", "Delivery lunch capture", "Agency", "Google", 12000, 10850, 41000, 11850, "Complete", "Medium", "Google Ads + aggregator report", **{"New Customers": 230, "Budget Owner Sign-off": "Yes"}),
                row("BUD-003", "Mar", "Content", "Founder video shoots", "Marketing Lead", "Instagram", 9000, 8400, 18000, 9600, "Live", "Medium", "Production invoices", **{"New Customers": 85, "Budget Owner Sign-off": "No"}),
                row("BUD-004", "Apr", "CRM", "Ramadan repeat visit flow", "Marketing Lead", "WhatsApp", 6500, 5200, 36500, 6100, "Complete", "High", "CRM export", **{"New Customers": 118, "Budget Owner Sign-off": "Yes"}),
                row("BUD-005", "May", "Local Store", "Mall sampling weekend", "Store Manager", "In-store", 7500, 9100, 22000, 10400, "At Risk", "High", "Store manager log", **{"New Customers": 145, "Budget Owner Sign-off": "No"}),
                row("BUD-006", "Jun", "Influencer", "Summer beverage tastings", "Agency", "Influencer", 11000, 9800, 29000, 11200, "Ready", "Medium", "Creator briefs", **{"New Customers": 190, "Budget Owner Sign-off": "No"}),
            ],
        ),
        WorkbookSpec(
            filename="annual-restaurant-marketing-plan.xlsx",
            folder="growth",
            title="Annual Restaurant Marketing Operating Plan",
            subtitle="Campaign calendar, seasonal moments, KPI impact, owner matrix, budget readiness, and board summary.",
            category="Annual operating plan",
            audience="Founders, marketing managers, franchise teams, and brand operators.",
            management_question="Does the annual plan connect business objectives, campaigns, owners, budgets, and expected KPI impact?",
            dimension_label="Business Objective",
            item_label="Campaign / Seasonal Moment",
            channel_label="Lead Channel",
            plan_label="Target KPI Lift",
            actual_label="Current Readiness %",
            value_label="Expected Revenue Impact AED",
            cost_label="Campaign Budget AED",
            efficiency_label="Expected Payoff",
            chart_title="Monthly readiness and campaign budget",
            gross_margin=0.60,
            min_roi=0.20,
            target_score=0.82,
            assumptions=[
                ("Annual campaign count target", 18, "campaigns", "Management target", "Sets operating cadence", "Marketing Lead"),
                ("Minimum readiness before launch", 0.80, "%", "Governance rule", "Blocks premature launches", "Operations Lead"),
                ("Seasonal planning lead time", 45, "days", "Operator benchmark", "Protects asset and approval timing", "Marketing Lead"),
                ("Board materiality threshold", 15000, "AED", "FP&A policy", "Requires commentary", "Finance Lead"),
            ],
            definitions=[
                ("Readiness score", "Weighted completion of owner, assets, offer, channel, and measurement setup.", "Average readiness inputs"),
                ("KPI lift", "Expected improvement in the commercial KPI the campaign is meant to move.", "Target KPI after campaign - baseline KPI"),
                ("Seasonal moment", "A market timing trigger such as Ramadan, summer, school return, national day, or mall event.", "N/A"),
                ("Owner matrix", "Clear accountability for strategy, assets, operations, finance, and approval.", "RACI-style ownership"),
            ],
            actions=[
                ("Low readiness", "Move campaigns below readiness threshold into a pre-launch blocker review.", "Operations Lead", "High"),
                ("Unclear objective", "Rewrite any campaign that cannot name the KPI it is meant to move.", "Marketing Lead", "High"),
                ("Budget concentration", "Check whether spend is over-weighted to one season or one channel.", "Finance Lead", "Medium"),
            ],
            extra_headers=["Seasonal Moment", "Offer / Hook", "Asset Readiness %", "Measurement Owner"],
            rows=[
                row("PLAN-001", "Jan", "Acquire new guests", "New Year trial offer", "Marketing Lead", "Meta", 0.12, 0.86, 68000, 17500, "Approved", "High", "Calendar + media plan", **{"Seasonal Moment": "New Year", "Offer / Hook": "Starter combo", "Asset Readiness %": 0.90, "Measurement Owner": "Finance Lead"}),
                row("PLAN-002", "Feb", "Lift delivery lunch", "Office lunch push", "Operations Lead", "Aggregator", 0.08, 0.72, 42000, 11500, "In Progress", "Medium", "Aggregator calendar", **{"Seasonal Moment": "Office routine", "Offer / Hook": "Lunch under 30 AED", "Asset Readiness %": 0.70, "Measurement Owner": "Marketing Lead"}),
                row("PLAN-003", "Mar", "Retain regulars", "Ramadan CRM flow", "Marketing Lead", "WhatsApp", 0.15, 0.91, 83000, 15500, "Ready", "Critical", "CRM briefs", **{"Seasonal Moment": "Ramadan", "Offer / Hook": "Repeat visit reward", "Asset Readiness %": 0.95, "Measurement Owner": "Marketing Lead"}),
                row("PLAN-004", "May", "Increase AOV", "Summer beverage bundles", "Store Manager", "In-store", 0.10, 0.64, 39000, 9800, "At Risk", "High", "Menu boards pending", **{"Seasonal Moment": "Summer", "Offer / Hook": "Cool drink add-on", "Asset Readiness %": 0.55, "Measurement Owner": "Store Manager"}),
                row("PLAN-005", "Sep", "Win school traffic", "Back to school snacks", "Marketing Lead", "OOH", 0.09, 0.58, 36000, 12500, "In Progress", "Medium", "Partner list", **{"Seasonal Moment": "Back to school", "Offer / Hook": "Family snack box", "Asset Readiness %": 0.45, "Measurement Owner": "Finance Lead"}),
                row("PLAN-006", "Dec", "Build brand heat", "National Day story", "Agency", "PR", 0.11, 0.50, 51000, 18000, "Not Started", "Medium", "Concept note", **{"Seasonal Moment": "National Day", "Offer / Hook": "Local pride story", "Asset Readiness %": 0.35, "Measurement Owner": "Marketing Lead"}),
            ],
        ),
        WorkbookSpec(
            filename="cafe-content-calendar.xlsx",
            folder="growth",
            title="Cafe Content Operations Calendar",
            subtitle="Content pillars, shoot tracker, approval flow, posting cadence, platform mix, and performance review.",
            category="Content operations",
            audience="Cafe owners, marketing teams, social media managers, and agencies.",
            management_question="Is content production disciplined enough to support campaigns, visibility, and sales without last-minute chaos?",
            dimension_label="Content Pillar",
            item_label="Content / Asset",
            channel_label="Platform",
            plan_label="Planned Assets",
            actual_label="Published Assets",
            value_label="Estimated Revenue Influence AED",
            cost_label="Production Cost AED",
            efficiency_label="Content Payoff",
            chart_title="Monthly planned vs published assets",
            gross_margin=0.58,
            min_roi=0.15,
            target_score=0.85,
            assumptions=[
                ("Minimum weekly posting cadence", 4, "posts", "Content system standard", "Protects brand freshness", "Marketing Lead"),
                ("Shoot lead time", 10, "days", "Production rule", "Prevents approval delays", "Agency"),
                ("Approval SLA", 48, "hours", "Governance rule", "Keeps calendar moving", "Founder"),
                ("Engagement review cadence", 7, "days", "Analytics rhythm", "Improves content mix", "Marketing Lead"),
            ],
            definitions=[
                ("Content pillar", "Recurring theme that helps the brand stay consistent across posts.", "N/A"),
                ("Posting cadence", "Planned publishing frequency by platform and week.", "Published Assets / Week"),
                ("Approval SLA", "Time between content submission and final approval.", "Approval date - submission date"),
                ("Content payoff", "Estimated gross profit influenced by content after production cost.", "=(Revenue influence x Gross Margin - Cost) / Cost"),
            ],
            actions=[
                ("Shoot backlog", "Batch menu, founder, and store assets in one monthly shoot day.", "Agency", "High"),
                ("Approval delay", "Set a 48-hour approval SLA with one final approver.", "Founder", "Medium"),
                ("Weak platform mix", "Shift low-performing generic posts toward product, proof, and store-level content.", "Marketing Lead", "Medium"),
            ],
            extra_headers=["Format", "Shoot Date", "Approval Stage", "Engagement Rate %", "Campaign Link"],
            rows=[
                row("CON-001", "Jan", "Product craving", "Signature drink reel", "Agency", "Instagram", 6, 5, 14500, 4200, "Complete", "High", "Published links", **{"Format": "Reel", "Shoot Date": date(2026, 1, 6), "Approval Stage": "Approved", "Engagement Rate %": 0.052, "Campaign Link": "Winter beverages"}),
                row("CON-002", "Jan", "Founder story", "Why this menu exists", "Founder", "TikTok", 3, 2, 7200, 1800, "Live", "Medium", "Draft + analytics", **{"Format": "Short video", "Shoot Date": date(2026, 1, 12), "Approval Stage": "Approved", "Engagement Rate %": 0.041, "Campaign Link": "Brand story"}),
                row("CON-003", "Feb", "Local community", "Office lunch carousel", "Marketing Lead", "Instagram", 4, 3, 9800, 2600, "In Progress", "Medium", "Design board", **{"Format": "Carousel", "Shoot Date": date(2026, 2, 4), "Approval Stage": "Review", "Engagement Rate %": 0.033, "Campaign Link": "Office lunch"}),
                row("CON-004", "Mar", "Offer support", "Ramadan bundle content", "Marketing Lead", "WhatsApp", 8, 4, 33500, 5200, "At Risk", "High", "Offer copy pending", **{"Format": "Stories + WhatsApp", "Shoot Date": date(2026, 2, 24), "Approval Stage": "Blocked", "Engagement Rate %": 0.029, "Campaign Link": "Ramadan"}),
                row("CON-005", "Apr", "Proof", "Customer review clips", "Store Manager", "TikTok", 5, 3, 11000, 2500, "Ready", "Medium", "Consent list", **{"Format": "UGC", "Shoot Date": date(2026, 4, 8), "Approval Stage": "Approved", "Engagement Rate %": 0.047, "Campaign Link": "Reviews"}),
                row("CON-006", "May", "Menu education", "How to order combo", "Agency", "Instagram", 4, 1, 5800, 2100, "Blocked", "High", "Menu pricing not final", **{"Format": "Reel", "Shoot Date": date(2026, 5, 3), "Approval Stage": "Blocked", "Engagement Rate %": 0.018, "Campaign Link": "Combo launch"}),
            ],
        ),
        WorkbookSpec(
            filename="crm-loyalty-campaign-planner.xlsx",
            folder="growth",
            title="CRM & Loyalty Growth Framework",
            subtitle="Segments, lifecycle campaigns, retention metrics, offer cost, redemption, customer value, and CRM ROI.",
            category="CRM and loyalty",
            audience="Operators, loyalty managers, CRM teams, and founders who need repeat revenue.",
            management_question="Are CRM and loyalty campaigns improving frequency, repeat rate, and customer value profitably?",
            dimension_label="Lifecycle Segment",
            item_label="Campaign / Journey",
            channel_label="CRM Channel",
            plan_label="Target Redemption / Response",
            actual_label="Actual Redemption / Response",
            value_label="Incremental Revenue AED",
            cost_label="Offer + Message Cost AED",
            efficiency_label="CRM ROI",
            chart_title="Monthly CRM cost vs incremental revenue",
            gross_margin=0.64,
            min_roi=0.30,
            target_score=0.85,
            assumptions=[
                ("Active member definition", 90, "days", "CRM rule", "Defines active base", "Marketing Lead"),
                ("Win-back lapse threshold", 45, "days", "Retention rule", "Triggers reactivation", "Marketing Lead"),
                ("Maximum offer food cost", 0.18, "% of order", "Finance guardrail", "Protects contribution", "Finance Lead"),
                ("Minimum redemption rate", 0.12, "%", "CRM benchmark", "Flags weak offer-market fit", "Marketing Lead"),
            ],
            definitions=[
                ("Repeat rate", "Share of customers who buy more than once in the period.", "=Repeat Customers / Total Customers"),
                ("Redemption rate", "Share of eligible customers who redeem the offer.", "=Redemptions / Eligible Audience"),
                ("Customer LTV", "Expected value from a customer over their relationship with the brand.", "=AOV x Frequency x Lifespan x Gross Margin"),
                ("CRM ROI", "Incremental gross profit less offer/message cost divided by cost.", "=(Incremental Revenue x Gross Margin - Cost) / Cost"),
            ],
            actions=[
                ("Low redemption", "Test simpler rewards and segment offers by visit history.", "Marketing Lead", "High"),
                ("Margin leakage", "Move rewards toward high-margin sides, drinks, or bundles.", "Finance Lead", "High"),
                ("Weak second visit", "Add a 48-hour post-first-order follow-up journey.", "Marketing Lead", "Medium"),
            ],
            extra_headers=["Audience Size", "Redemptions", "Repeat Rate %", "AOV AED", "Customer LTV AED"],
            rows=[
                row("CRM-001", "Jan", "First-time guests", "Second visit nudge", "Marketing Lead", "WhatsApp", 0.14, 0.16, 27500, 4200, "Complete", "High", "CRM export", **{"Audience Size": 1800, "Redemptions": 288, "Repeat Rate %": 0.22, "AOV AED": 44, "Customer LTV AED": 214}),
                row("CRM-002", "Feb", "Regulars", "VIP drink add-on", "Marketing Lead", "SMS", 0.18, 0.15, 19500, 3800, "Complete", "Medium", "SMS report", **{"Audience Size": 920, "Redemptions": 138, "Repeat Rate %": 0.41, "AOV AED": 51, "Customer LTV AED": 462}),
                row("CRM-003", "Mar", "Lapsed 45 days", "Win-back ladder 1", "Marketing Lead", "WhatsApp", 0.12, 0.09, 14200, 3900, "At Risk", "High", "Win-back cohort", **{"Audience Size": 1300, "Redemptions": 117, "Repeat Rate %": 0.12, "AOV AED": 39, "Customer LTV AED": 136}),
                row("CRM-004", "Apr", "Birthday", "Birthday dessert reward", "Store Manager", "Email", 0.20, 0.23, 11800, 1600, "Live", "Medium", "Automation log", **{"Audience Size": 240, "Redemptions": 55, "Repeat Rate %": 0.36, "AOV AED": 57, "Customer LTV AED": 325}),
                row("CRM-005", "May", "Delivery customers", "Direct ordering switch", "Marketing Lead", "Push", 0.08, 0.05, 8900, 3100, "Blocked", "High", "App push not ready", **{"Audience Size": 2100, "Redemptions": 105, "Repeat Rate %": 0.08, "AOV AED": 42, "Customer LTV AED": 128}),
                row("CRM-006", "Jun", "High AOV guests", "Private tasting invite", "Founder", "Email", 0.10, 0.12, 36000, 6200, "Ready", "Medium", "Invite list", **{"Audience Size": 300, "Redemptions": 36, "Repeat Rate %": 0.48, "AOV AED": 96, "Customer LTV AED": 880}),
            ],
        ),
        WorkbookSpec(
            filename="delivery-aggregator-audit.xlsx",
            folder="growth",
            title="Delivery Aggregator Audit & Profitability Model",
            subtitle="Menu visibility, pricing, imagery, ratings, delivery time, commission impact, discount dependency, and action priority.",
            category="Delivery audit",
            audience="Delivery-first brands, cloud kitchens, multi-channel stores, and operators.",
            management_question="Is delivery volume profitable, visible, and operationally reliable enough to scale?",
            dimension_label="Platform / Store",
            item_label="Audit Area / Menu Item",
            channel_label="Aggregator",
            plan_label="Target Score / Benchmark",
            actual_label="Current Score / Actual",
            value_label="Monthly Delivery Revenue AED",
            cost_label="Commission + Promo Cost AED",
            efficiency_label="Net Delivery ROI",
            chart_title="Delivery revenue vs platform cost",
            gross_margin=0.55,
            min_roi=0.10,
            target_score=0.82,
            assumptions=[
                ("Commission rate watch band", 0.28, "%", "Aggregator contract review", "Flags margin compression", "Finance Lead"),
                ("Rating risk threshold", 4.2, "stars", "Platform benchmark", "Impacts conversion", "Operations Lead"),
                ("Prep time target", 18, "minutes", "Ops standard", "Protects ranking and guest experience", "Store Manager"),
                ("Photo coverage target", 0.90, "% of menu", "Delivery merchandising rule", "Improves conversion", "Marketing Lead"),
            ],
            definitions=[
                ("Commission impact", "Platform commission and promo funding as a share of delivery revenue.", "=Commission + Promo Cost / Delivery Revenue"),
                ("Discount dependency", "Revenue share that requires funded discounts to convert.", "=Discounted Orders / Total Orders"),
                ("Net delivery ROI", "Gross profit after platform costs divided by platform costs.", "=(Delivery Revenue x Gross Margin - Platform Cost) / Platform Cost"),
                ("Ranking risk", "Operational and merchandising issues that can reduce visibility or conversion.", "Scorecard threshold test"),
            ],
            actions=[
                ("Negative contribution", "Remove or reprice items that cannot survive commission and packaging cost.", "Finance Lead", "Critical"),
                ("Weak menu visuals", "Prioritize photos and descriptions for top 20 selling items.", "Marketing Lead", "High"),
                ("Slow prep time", "Separate delivery mise en place and reduce fragile items from aggregator menu.", "Operations Lead", "High"),
            ],
            extra_headers=["Commission %", "Rating", "Avg Prep Time Min", "Photo Coverage %", "Discount Dependency %"],
            rows=[
                row("DEL-001", "Jan", "Deliveroo - Marina", "Top item photo coverage", "Marketing Lead", "Aggregator", 0.90, 0.76, 88500, 26400, "At Risk", "High", "Platform audit", **{"Commission %": 0.28, "Rating": 4.3, "Avg Prep Time Min": 21, "Photo Coverage %": 0.76, "Discount Dependency %": 0.31}),
                row("DEL-002", "Feb", "Talabat - Downtown", "Hero combo margin", "Finance Lead", "Aggregator", 0.25, 0.18, 64000, 19800, "Blocked", "Critical", "Payout report", **{"Commission %": 0.30, "Rating": 4.1, "Avg Prep Time Min": 24, "Photo Coverage %": 0.68, "Discount Dependency %": 0.42}),
                row("DEL-003", "Mar", "Noon - JLT", "Menu descriptions", "Marketing Lead", "Aggregator", 0.85, 0.82, 42000, 10400, "Ready", "Medium", "Listing review", **{"Commission %": 0.24, "Rating": 4.4, "Avg Prep Time Min": 17, "Photo Coverage %": 0.82, "Discount Dependency %": 0.22}),
                row("DEL-004", "Apr", "Careem - Business Bay", "Delivery-only bundle", "Operations Lead", "Aggregator", 0.30, 0.34, 38000, 9200, "Live", "Medium", "Bundle report", **{"Commission %": 0.23, "Rating": 4.5, "Avg Prep Time Min": 16, "Photo Coverage %": 0.88, "Discount Dependency %": 0.18}),
                row("DEL-005", "May", "Talabat - Sharjah", "Ratings recovery", "Store Manager", "Aggregator", 4.3, 4.0, 29000, 8100, "At Risk", "High", "Review export", **{"Commission %": 0.27, "Rating": 4.0, "Avg Prep Time Min": 22, "Photo Coverage %": 0.72, "Discount Dependency %": 0.35}),
                row("DEL-006", "Jun", "Direct web", "Aggregator-to-direct insert", "Marketing Lead", "Email", 0.12, 0.08, 15500, 2100, "In Progress", "High", "Bag insert test", **{"Commission %": 0.04, "Rating": 4.6, "Avg Prep Time Min": 15, "Photo Coverage %": 0.95, "Discount Dependency %": 0.10}),
            ],
        ),
        WorkbookSpec(
            filename="franchise-campaign-rollout-tracker.xlsx",
            folder="growth",
            title="Franchise Campaign Rollout Governance Tracker",
            subtitle="Store readiness, compliance, asset deployment, local adaptation, approvals, launch score, and issue log.",
            category="Franchise rollout",
            audience="Franchise teams, field consultants, HQ marketers, and multi-unit operators.",
            management_question="Can this campaign launch consistently across stores without brand, compliance, or execution leakage?",
            dimension_label="Store / Franchisee",
            item_label="Rollout Requirement",
            channel_label="Market / Region",
            plan_label="Target Compliance %",
            actual_label="Current Compliance %",
            value_label="Expected Sales Lift AED",
            cost_label="Rollout Support Cost AED",
            efficiency_label="Rollout ROI",
            chart_title="Store compliance and rollout cost",
            gross_margin=0.60,
            min_roi=0.18,
            target_score=0.88,
            assumptions=[
                ("Minimum go-live compliance", 0.85, "%", "Franchise governance rule", "Blocks weak stores", "Operations Lead"),
                ("Asset deployment SLA", 7, "days before launch", "Campaign SOP", "Protects local setup", "Marketing Lead"),
                ("Issue closure SLA", 5, "business days", "Field support rule", "Controls rollout risk", "Operations Lead"),
                ("Post-campaign report due", 10, "days after end", "Review cadence", "Captures learnings", "Finance Lead"),
            ],
            definitions=[
                ("Launch score", "Weighted readiness across assets, approvals, training, local setup, and measurement.", "Average scored requirements"),
                ("Compliance score", "Share of campaign standards completed by each store.", "=Completed Requirements / Total Requirements"),
                ("Local adaptation", "Approved local changes that preserve central brand standards.", "Approval log"),
                ("Issue closure rate", "Share of rollout issues closed within SLA.", "=Closed on time / Total issues"),
            ],
            actions=[
                ("Store readiness gap", "Escalate stores below 85% compliance before central launch approval.", "Operations Lead", "High"),
                ("Missing local approvals", "Create one approval gate for local adaptations and offers.", "Marketing Lead", "High"),
                ("Weak reporting", "Require post-campaign sales and compliance report from every franchisee.", "Finance Lead", "Medium"),
            ],
            extra_headers=["Assets Ready %", "Training Complete %", "Local Approval", "Issue Count", "Launch Score %"],
            rows=[
                row("FR-001", "Jan", "Dubai Mall", "Window decals + menu boards", "Franchisee", "Dubai", 0.90, 0.88, 52000, 7400, "Ready", "High", "Field audit", **{"Assets Ready %": 0.95, "Training Complete %": 0.82, "Local Approval": "Yes", "Issue Count": 1, "Launch Score %": 0.88}),
                row("FR-002", "Feb", "Abu Dhabi Corniche", "Staff briefing and offer setup", "Store Manager", "Abu Dhabi", 0.90, 0.74, 36000, 6400, "At Risk", "High", "Training log", **{"Assets Ready %": 0.80, "Training Complete %": 0.68, "Local Approval": "No", "Issue Count": 4, "Launch Score %": 0.72}),
                row("FR-003", "Mar", "Sharjah City Centre", "Aggregator update", "Franchisee", "Sharjah", 0.85, 0.91, 29500, 3100, "Complete", "Medium", "Platform screenshots", **{"Assets Ready %": 0.93, "Training Complete %": 0.90, "Local Approval": "Yes", "Issue Count": 0, "Launch Score %": 0.91}),
                row("FR-004", "Apr", "Riyadh North", "Arabic asset localization", "Marketing Lead", "Saudi", 0.88, 0.62, 67000, 11500, "Blocked", "Critical", "Translation pending", **{"Assets Ready %": 0.58, "Training Complete %": 0.60, "Local Approval": "No", "Issue Count": 6, "Launch Score %": 0.61}),
                row("FR-005", "May", "Doha Pearl", "Local influencer launch", "Agency", "Qatar", 0.80, 0.79, 41000, 8800, "In Progress", "Medium", "Creator list", **{"Assets Ready %": 0.82, "Training Complete %": 0.75, "Local Approval": "Yes", "Issue Count": 2, "Launch Score %": 0.79}),
                row("FR-006", "Jun", "Jeddah Waterfront", "Campaign compliance pack", "Operations Lead", "Saudi", 0.88, 0.86, 53000, 7100, "Approved", "High", "Approval checklist", **{"Assets Ready %": 0.90, "Training Complete %": 0.84, "Local Approval": "Yes", "Issue Count": 1, "Launch Score %": 0.86}),
            ],
        ),
        WorkbookSpec(
            filename="google-business-profile-checklist.xlsx",
            folder="growth",
            title="Google Business Profile & Local SEO Audit",
            subtitle="Profile completeness, reviews, response SLA, photo freshness, keyword signals, competitor benchmark, and priority actions.",
            category="Local SEO audit",
            audience="Independent restaurants, local store marketers, and multi-location teams.",
            management_question="Is each store easy to find, trust, and choose on Google Search and Maps?",
            dimension_label="Store / Listing",
            item_label="GBP Audit Area",
            channel_label="Local Search Surface",
            plan_label="Target Local Score",
            actual_label="Current Local Score",
            value_label="Estimated Local Demand AED",
            cost_label="Fix Cost / Effort AED",
            efficiency_label="Local SEO Payoff",
            chart_title="Local score vs estimated demand",
            gross_margin=0.60,
            min_roi=0.20,
            target_score=0.90,
            assumptions=[
                ("Review response SLA", 48, "hours", "Local SEO governance", "Improves trust and service recovery", "Store Manager"),
                ("Photo freshness target", 30, "days", "Profile freshness rule", "Signals active store", "Marketing Lead"),
                ("Minimum rating target", 4.4, "stars", "Reputation guardrail", "Protects conversion", "Operations Lead"),
                ("Post frequency target", 4, "posts / month", "GBP cadence", "Supports freshness", "Marketing Lead"),
            ],
            definitions=[
                ("Profile completeness", "Share of key GBP fields completed accurately.", "=Completed fields / Required fields"),
                ("Review SLA", "Average time to respond to customer reviews.", "Response time in hours"),
                ("Prominence", "Google local signal linked to review count, rating, links, and broader web presence.", "Qualitative + quantitative benchmark"),
                ("Local action value", "Estimated value from calls, website clicks, directions, and search visibility.", "Actions x conversion x AOV"),
            ],
            actions=[
                ("Incomplete profile", "Fix hours, attributes, menu, ordering link, services, and business description.", "Marketing Lead", "Critical"),
                ("Slow review replies", "Assign daily review owner and response templates by sentiment.", "Store Manager", "High"),
                ("Stale photos", "Upload recent food, storefront, menu, and team photos every month.", "Marketing Lead", "Medium"),
            ],
            extra_headers=["Rating", "Review Response Hrs", "Photo Age Days", "Posts / Month", "Competitor Gap"],
            rows=[
                row("GBP-001", "Jan", "Dubai Marina", "Profile completeness", "Marketing Lead", "Google Maps", 0.95, 0.88, 18500, 1200, "In Progress", "High", "GBP dashboard", **{"Rating": 4.5, "Review Response Hrs": 38, "Photo Age Days": 18, "Posts / Month": 3, "Competitor Gap": "Photos"}),
                row("GBP-002", "Feb", "Downtown", "Review response SLA", "Store Manager", "Google Search", 0.90, 0.66, 22500, 900, "At Risk", "Critical", "Review export", **{"Rating": 4.2, "Review Response Hrs": 96, "Photo Age Days": 44, "Posts / Month": 1, "Competitor Gap": "Reviews"}),
                row("GBP-003", "Mar", "JLT", "Menu and ordering links", "Marketing Lead", "Google Maps", 0.92, 0.91, 14800, 650, "Complete", "Medium", "Listing screenshot", **{"Rating": 4.6, "Review Response Hrs": 22, "Photo Age Days": 12, "Posts / Month": 4, "Competitor Gap": "Low"}),
                row("GBP-004", "Apr", "Sharjah", "Photo freshness", "Agency", "Google Maps", 0.85, 0.58, 9800, 1500, "Blocked", "High", "Photo bank missing", **{"Rating": 4.1, "Review Response Hrs": 72, "Photo Age Days": 90, "Posts / Month": 0, "Competitor Gap": "Photos + rating"}),
                row("GBP-005", "May", "Abu Dhabi", "Keyword alignment", "Marketing Lead", "Google Search", 0.80, 0.76, 17300, 1100, "Ready", "Medium", "Landing page review", **{"Rating": 4.4, "Review Response Hrs": 41, "Photo Age Days": 24, "Posts / Month": 2, "Competitor Gap": "Categories"}),
                row("GBP-006", "Jun", "Riyadh", "Competitor benchmark", "Operations Lead", "Google Maps", 0.82, 0.69, 26000, 2200, "In Progress", "High", "Map pack export", **{"Rating": 4.3, "Review Response Hrs": 55, "Photo Age Days": 36, "Posts / Month": 2, "Competitor Gap": "Review count"}),
            ],
        ),
        WorkbookSpec(
            filename="local-store-marketing-planner.xlsx",
            folder="growth",
            title="Local Store Marketing Strategy Planner",
            subtitle="Trade-area mapping, local segments, partnerships, hyperlocal campaigns, footfall estimates, budget, and ROI tracking.",
            category="Local store marketing",
            audience="Store managers, local marketers, franchisees, and single-outlet founders.",
            management_question="Which local moves will bring realistic footfall and repeat visits for this exact trade area?",
            dimension_label="Trade Area Segment",
            item_label="Local Activity / Partner",
            channel_label="Activation Channel",
            plan_label="Target Footfall / Leads",
            actual_label="Actual Footfall / Leads",
            value_label="Estimated Sales Impact AED",
            cost_label="Local Marketing Cost AED",
            efficiency_label="Local ROI",
            chart_title="Local footfall plan vs actual",
            gross_margin=0.61,
            min_roi=0.20,
            target_score=0.82,
            assumptions=[
                ("Trade area radius", 3, "km", "Local planning rule", "Defines target clusters", "Store Manager"),
                ("Conversion from sample to visit", 0.18, "%", "Operator assumption", "Estimates local ROI", "Marketing Lead"),
                ("Office partnership minimum", 5, "partners", "Local playbook", "Builds weekday demand", "Store Manager"),
                ("Local campaign review cadence", 14, "days", "Management rhythm", "Stops dead activities", "Marketing Lead"),
            ],
            definitions=[
                ("Trade area", "The practical local catchment where customers can visit or order from the store.", "Radius + demand clusters"),
                ("Footfall estimate", "Expected store visits generated by a local activity.", "Reach x conversion rate"),
                ("Partnership value", "Sales impact from schools, offices, residential communities, gyms, malls, or local businesses.", "Estimated orders x AOV"),
                ("Local ROI", "Gross profit from local sales impact less activity cost, divided by cost.", "=(Sales impact x Gross Margin - Cost) / Cost"),
            ],
            actions=[
                ("Weak office pipeline", "Build a target list of nearby offices with decision-maker, offer, and follow-up date.", "Store Manager", "High"),
                ("Low conversion", "Change offer from generic discount to specific trial reason by segment.", "Marketing Lead", "Medium"),
                ("Untracked activities", "Require QR, coupon code, or POS tag for every local activation.", "Finance Lead", "High"),
            ],
            extra_headers=["Cluster Type", "Distance Km", "Partner Count", "Conversion Rate %", "AOV AED"],
            rows=[
                row("LSM-001", "Jan", "Office cluster", "Lunch tasting at business tower", "Store Manager", "In-store", 220, 185, 13800, 3100, "Complete", "High", "Sampling log", **{"Cluster Type": "Office", "Distance Km": 0.7, "Partner Count": 3, "Conversion Rate %": 0.16, "AOV AED": 42}),
                row("LSM-002", "Feb", "Residential", "Weekend family bundle flyer", "Marketing Lead", "Flyers", 160, 118, 7200, 1800, "Live", "Medium", "Coupon code", **{"Cluster Type": "Residential", "Distance Km": 1.5, "Partner Count": 1, "Conversion Rate %": 0.09, "AOV AED": 61}),
                row("LSM-003", "Mar", "School", "After-school snack tie-up", "Store Manager", "OOH", 190, 142, 8600, 2600, "In Progress", "High", "School approvals", **{"Cluster Type": "School", "Distance Km": 1.1, "Partner Count": 2, "Conversion Rate %": 0.12, "AOV AED": 37}),
                row("LSM-004", "Apr", "Gym", "Protein coffee collaboration", "Founder", "PR", 100, 125, 11200, 2400, "Complete", "Medium", "Partner report", **{"Cluster Type": "Fitness", "Distance Km": 0.5, "Partner Count": 1, "Conversion Rate %": 0.22, "AOV AED": 48}),
                row("LSM-005", "May", "Mall traffic", "Entrance sampling", "Operations Lead", "In-store", 300, 165, 9900, 4200, "At Risk", "High", "Mall permit issue", **{"Cluster Type": "Mall", "Distance Km": 0.0, "Partner Count": 1, "Conversion Rate %": 0.07, "AOV AED": 45}),
                row("LSM-006", "Jun", "Hotel nearby", "Concierge card", "Marketing Lead", "PR", 80, 54, 6800, 900, "Ready", "Low", "Hotel list", **{"Cluster Type": "Hotel", "Distance Km": 0.8, "Partner Count": 4, "Conversion Rate %": 0.11, "AOV AED": 63}),
            ],
        ),
        WorkbookSpec(
            filename="marketing-roi-calculator.xlsx",
            folder="growth",
            title="Campaign ROI & Profitability Calculator",
            subtitle="Spend, impressions, clicks, conversions, CAC, gross profit, contribution margin, ROAS, payback, breakeven, and sensitivity.",
            category="Marketing ROI",
            audience="Founders, marketers, finance reviewers, and operators approving campaign spend.",
            management_question="Did the campaign create profitable incremental sales, or just activity and platform-reported revenue?",
            dimension_label="Campaign Type",
            item_label="Campaign Name",
            channel_label="Channel",
            plan_label="Planned Conversions",
            actual_label="Actual Conversions",
            value_label="Attributed Revenue AED",
            cost_label="Total Campaign Cost AED",
            efficiency_label="Profit ROI",
            chart_title="Campaign cost vs attributed revenue",
            gross_margin=0.60,
            min_roi=0.25,
            target_score=0.85,
            assumptions=[
                ("Attribution confidence factor", 0.75, "%", "Analyst assumption", "Discounts inflated platform revenue", "Marketing Lead"),
                ("Payback target", 2, "months", "Management target", "Flags slow payback", "Finance Lead"),
                ("Breakeven ROAS", 1.7, "x", "Gross margin logic", "Minimum revenue multiple", "Finance Lead"),
                ("Incrementality haircut", 0.20, "%", "Conservative review", "Avoids over-crediting marketing", "Finance Lead"),
            ],
            definitions=[
                ("ROAS", "Revenue attributed to campaign divided by ad or campaign spend.", "=Revenue / Spend"),
                ("Profit ROI", "Gross profit less total campaign cost divided by total campaign cost.", "=(Revenue x Gross Margin - Cost) / Cost"),
                ("CAC", "Total campaign cost divided by new customers acquired.", "=Cost / New Customers"),
                ("Payback", "Months required for gross profit to recover campaign cost.", "=Cost / Monthly Gross Profit"),
            ],
            actions=[
                ("ROAS without margin", "Report both ROAS and profit ROI before scaling spend.", "Finance Lead", "High"),
                ("Poor conversion", "Review offer, landing page, and tracking before increasing budget.", "Marketing Lead", "High"),
                ("Unclear incrementality", "Apply an attribution confidence factor or holdout test for major campaigns.", "Finance Lead", "Medium"),
            ],
            extra_headers=["Impressions", "Clicks", "CTR %", "New Customers", "CAC AED", "ROAS x", "Payback Months"],
            rows=[
                row("ROI-001", "Jan", "Paid Social", "Trial combo launch", "Marketing Lead", "Meta", 520, 585, 68000, 18200, "Complete", "High", "Ads + POS tag", **{"Impressions": 210000, "Clicks": 6100, "CTR %": 0.029, "New Customers": 420}),
                row("ROI-002", "Feb", "Search", "Near me lunch ads", "Agency", "Google", 210, 188, 31500, 9200, "Complete", "Medium", "Google Ads", **{"Impressions": 56000, "Clicks": 2450, "CTR %": 0.044, "New Customers": 160}),
                row("ROI-003", "Mar", "CRM", "Win-back offer", "Marketing Lead", "WhatsApp", 150, 132, 24400, 4100, "Live", "High", "CRM export", **{"Impressions": 9000, "Clicks": 1170, "CTR %": 0.130, "New Customers": 48}),
                row("ROI-004", "Apr", "Influencer", "Creator tasting week", "Agency", "Influencer", 220, 95, 17800, 13200, "At Risk", "High", "Creator codes", **{"Impressions": 320000, "Clicks": 3800, "CTR %": 0.012, "New Customers": 70}),
                row("ROI-005", "May", "Local", "Mall sampling", "Store Manager", "In-store", 180, 214, 19800, 6500, "Complete", "Medium", "Coupon count", **{"Impressions": 12000, "Clicks": 0, "CTR %": 0, "New Customers": 126}),
                row("ROI-006", "Jun", "Aggregator", "BOGO listing push", "Operations Lead", "Aggregator", 300, 340, 41500, 18800, "Ready", "Critical", "Aggregator promo forecast", **{"Impressions": 88000, "Clicks": 5400, "CTR %": 0.061, "New Customers": 230}),
            ],
        ),
        WorkbookSpec(
            filename="menu-launch-and-offer-planner.xlsx",
            folder="growth",
            title="Menu Launch & Offer Profitability Model",
            subtitle="Item cost, selling price, gross margin, forecast volume, launch cost, breakeven, cannibalization, and readiness.",
            category="Menu and offers",
            audience="Chefs, founders, operations teams, and marketing managers launching menu items or offers.",
            management_question="Will this item or offer add profitable demand, or will it create volume with weak margin?",
            dimension_label="Menu Category",
            item_label="Item / Offer",
            channel_label="Launch Channel",
            plan_label="Forecast Units",
            actual_label="Committed / Test Units",
            value_label="Forecast Sales AED",
            cost_label="Launch + Promo Cost AED",
            efficiency_label="Offer ROI",
            chart_title="Forecast units and launch cost",
            gross_margin=0.65,
            min_roi=0.22,
            target_score=0.84,
            assumptions=[
                ("Target item gross margin", 0.65, "%", "Menu engineering target", "Protects contribution", "Finance Lead"),
                ("Cannibalization watch band", 0.20, "%", "Commercial guardrail", "Flags weak incrementality", "Marketing Lead"),
                ("Launch readiness threshold", 0.85, "%", "Go-live rule", "Protects execution quality", "Operations Lead"),
                ("Food cost update cadence", 30, "days", "Procurement rule", "Keeps margin current", "Operations Lead"),
            ],
            definitions=[
                ("Food cost %", "Ingredient cost divided by selling price.", "=Food Cost / Selling Price"),
                ("Gross margin %", "Selling price less food cost, divided by selling price.", "=(Selling Price - Food Cost) / Selling Price"),
                ("Breakeven units", "Units needed to recover launch and promotion cost.", "=Launch Cost / Contribution per Unit"),
                ("Cannibalization", "Sales taken from existing items instead of incremental demand.", "Cannibalized Sales / Forecast Sales"),
            ],
            actions=[
                ("Low margin item", "Reprice, reformulate, or bundle with high-margin add-ons before launch.", "Finance Lead", "Critical"),
                ("High cannibalization", "Position offer for a new occasion, daypart, or customer segment.", "Marketing Lead", "High"),
                ("Readiness gap", "Block launch until menu boards, staff scripts, photos, and POS setup are complete.", "Operations Lead", "High"),
            ],
            extra_headers=["Food Cost AED", "Selling Price AED", "Gross Margin %", "Breakeven Units", "Cannibalization %"],
            rows=[
                row("MENU-001", "Jan", "Beverage", "Saffron cold coffee", "Founder", "In-store", 1200, 950, 28500, 6200, "Ready", "High", "Recipe card", **{"Food Cost AED": 8.2, "Selling Price AED": 30, "Cannibalization %": 0.08}),
                row("MENU-002", "Feb", "Food", "Loaded paratha wrap", "Operations Lead", "Aggregator", 850, 630, 22050, 7400, "At Risk", "High", "Trial batch", **{"Food Cost AED": 13.8, "Selling Price AED": 35, "Cannibalization %": 0.24}),
                row("MENU-003", "Mar", "Bundle", "Ramadan iftar box", "Marketing Lead", "WhatsApp", 1500, 1320, 79200, 14500, "Approved", "Critical", "Bundle costing", **{"Food Cost AED": 26.5, "Selling Price AED": 60, "Cannibalization %": 0.12}),
                row("MENU-004", "Apr", "Dessert", "Pistachio mini cup", "Chef", "In-store", 700, 480, 9600, 2700, "In Progress", "Medium", "Supplier quote", **{"Food Cost AED": 5.7, "Selling Price AED": 20, "Cannibalization %": 0.05}),
                row("MENU-005", "May", "Offer", "Buy 2 chai get pastry", "Marketing Lead", "Meta", 1100, 860, 30100, 9200, "Blocked", "High", "POS setup pending", **{"Food Cost AED": 14.0, "Selling Price AED": 35, "Cannibalization %": 0.30}),
                row("MENU-006", "Jun", "Beverage", "Summer mango cooler", "Store Manager", "In-store", 1350, 1025, 32800, 5800, "Live", "High", "Menu boards", **{"Food Cost AED": 7.5, "Selling Price AED": 32, "Cannibalization %": 0.09}),
            ],
        ),
        WorkbookSpec(
            filename="restaurant-brand-positioning-brief.xlsx",
            folder="growth",
            title="Restaurant Brand Positioning & Investor Clarity Brief",
            subtitle="Brand canvas, personas, competitor comparison, differentiation score, proof points, messaging hierarchy, and clarity score.",
            category="Brand strategy",
            audience="Founders, brand teams, franchise prospects, and investors reviewing concept clarity.",
            management_question="Can the brand be understood, defended, repeated, and scaled without confusing customers or partners?",
            dimension_label="Strategy Area",
            item_label="Positioning Element",
            channel_label="Customer Touchpoint",
            plan_label="Target Clarity Score",
            actual_label="Current Clarity Score",
            value_label="Strategic Revenue Potential AED",
            cost_label="Brand Work Cost AED",
            efficiency_label="Brand Clarity Payoff",
            chart_title="Brand clarity by strategy area",
            gross_margin=0.62,
            min_roi=0.10,
            target_score=0.88,
            assumptions=[
                ("Minimum investor clarity score", 0.85, "%", "Board-ready threshold", "Supports scale story", "Founder"),
                ("Competitor set size", 5, "brands", "Strategy benchmark", "Improves positioning quality", "Marketing Lead"),
                ("Proof point requirement", 3, "proof points", "Messaging rule", "Stops empty claims", "Founder"),
                ("Persona review cadence", 180, "days", "Strategy refresh", "Keeps customer insight current", "Marketing Lead"),
            ],
            definitions=[
                ("Positioning", "The clear place the brand owns in the customer's mind and market.", "One-sentence test"),
                ("Differentiation score", "How hard the brand's advantage is to copy.", "Weighted score across product, story, proof, channel, operation"),
                ("Proof point", "Specific evidence that makes the brand promise believable.", "Claim + evidence"),
                ("Messaging hierarchy", "Order of message: who it is for, why it matters, what makes it different, and proof.", "Narrative structure"),
            ],
            actions=[
                ("Weak one-liner", "Rewrite positioning until a non-marketer can repeat it in one sentence.", "Founder", "High"),
                ("Thin proof", "Add operational, customer, product, and founder proof points to every claim.", "Marketing Lead", "High"),
                ("Copycat risk", "Map competitors and sharpen the area that is hardest to copy.", "Founder", "Medium"),
            ],
            extra_headers=["Persona", "Competitor Gap", "Proof Points Count", "One Sentence Ready", "Differentiation Score %"],
            rows=[
                row("BRAND-001", "Jan", "Customer", "Busy office worker persona", "Founder", "Website", 0.90, 0.74, 42000, 5500, "In Progress", "High", "Persona notes", **{"Persona": "Office Regular", "Competitor Gap": "Speed + routine", "Proof Points Count": 2, "One Sentence Ready": "No", "Differentiation Score %": 0.68}),
                row("BRAND-002", "Feb", "Promise", "Everyday premium chai", "Founder", "Store", 0.92, 0.86, 51000, 3200, "Approved", "High", "Brand canvas", **{"Persona": "Daily Treat", "Competitor Gap": "Consistency", "Proof Points Count": 4, "One Sentence Ready": "Yes", "Differentiation Score %": 0.81}),
                row("BRAND-003", "Mar", "Competitors", "Cafe chain comparison", "Marketing Lead", "Investor deck", 0.85, 0.62, 35000, 4800, "At Risk", "Medium", "Competitor table", **{"Persona": "Investor", "Competitor Gap": "Not quantified", "Proof Points Count": 1, "One Sentence Ready": "No", "Differentiation Score %": 0.54}),
                row("BRAND-004", "Apr", "Product cue", "Signature product naming", "Marketing Lead", "Menu", 0.80, 0.78, 26000, 2700, "Ready", "Medium", "Menu draft", **{"Persona": "New Guest", "Competitor Gap": "Product memory", "Proof Points Count": 3, "One Sentence Ready": "Yes", "Differentiation Score %": 0.76}),
                row("BRAND-005", "May", "Store cue", "What the store must signal", "Operations Lead", "Store", 0.88, 0.58, 68000, 12000, "Blocked", "High", "Design review", **{"Persona": "Walk-in Guest", "Competitor Gap": "Experience", "Proof Points Count": 2, "One Sentence Ready": "No", "Differentiation Score %": 0.49}),
                row("BRAND-006", "Jun", "Messaging", "Investor-level story", "Founder", "Investor deck", 0.90, 0.82, 75000, 8500, "In Progress", "High", "Deck outline", **{"Persona": "Investor", "Competitor Gap": "Scale logic", "Proof Points Count": 4, "One Sentence Ready": "Yes", "Differentiation Score %": 0.80}),
            ],
        ),
        WorkbookSpec(
            filename="restaurant-kpi-dashboard-template.xlsx",
            folder="growth",
            title="Restaurant Performance KPI Dashboard",
            subtitle="Weekly/monthly sales, footfall, AOV, delivery mix, repeat rate, labor cost, food cost, margin, marketing, ROAS, and alerts.",
            category="Performance dashboard",
            audience="Operators, GMs, finance leads, investors, and restaurant growth teams.",
            management_question="What is the business performance, risk, variance, and next management action?",
            dimension_label="Store / KPI Area",
            item_label="KPI Line Item",
            channel_label="Trading Channel",
            plan_label="Target KPI Value",
            actual_label="Actual KPI Value",
            value_label="Net Sales AED",
            cost_label="Variable Cost AED",
            efficiency_label="KPI Health",
            chart_title="Net sales and cost trend",
            gross_margin=0.63,
            min_roi=0.20,
            target_score=0.85,
            assumptions=[
                ("Prime cost watch band", 0.65, "% of sales", "Restaurant finance rule", "Flags margin pressure", "Finance Lead"),
                ("Food cost target", 0.32, "% of sales", "Kitchen benchmark", "Controls menu profitability", "Operations Lead"),
                ("Labor cost target", 0.28, "% of sales", "Scheduling target", "Controls productivity", "Operations Lead"),
                ("Repeat rate target", 0.35, "%", "Retention target", "Protects growth quality", "Marketing Lead"),
            ],
            definitions=[
                ("Prime cost", "Food cost plus labor cost as a percentage of sales.", "=(Food Cost + Labor Cost) / Sales"),
                ("AOV / APC", "Average order value or average spend per cover.", "=Net Sales / Covers"),
                ("Delivery mix", "Delivery sales as a share of total sales.", "=Delivery Sales / Net Sales"),
                ("ROAS", "Revenue attributed to marketing divided by marketing spend.", "=Attributed Revenue / Marketing Spend"),
            ],
            actions=[
                ("Prime cost pressure", "Review food cost, labor scheduling, waste, and discount leakage by store.", "Finance Lead", "Critical"),
                ("Traffic decline", "Separate guest count drop from AOV growth to avoid false comfort.", "Operations Lead", "High"),
                ("Marketing inefficiency", "Tie marketing spend to revenue contribution and repeat rate, not reach.", "Marketing Lead", "High"),
            ],
            extra_headers=["Covers", "AOV AED", "Delivery Mix %", "Food Cost %", "Labor Cost %", "Prime Cost %"],
            rows=[
                row("KPI-001", "Jan", "Store A", "Weekly net sales", "Store Manager", "All", 185000, 192400, 192400, 71000, "Complete", "High", "POS export", **{"Covers": 4680, "AOV AED": 41.1, "Delivery Mix %": 0.34, "Food Cost %": 0.31, "Labor Cost %": 0.27, "Prime Cost %": 0.58}),
                row("KPI-002", "Feb", "Store B", "Traffic recovery", "Operations Lead", "Dine-in", 142000, 131500, 131500, 58500, "At Risk", "High", "POS export", **{"Covers": 3010, "AOV AED": 43.7, "Delivery Mix %": 0.22, "Food Cost %": 0.34, "Labor Cost %": 0.31, "Prime Cost %": 0.65}),
                row("KPI-003", "Mar", "Store C", "Delivery mix review", "Marketing Lead", "Aggregator", 98000, 106700, 106700, 51400, "Live", "Medium", "Aggregator + POS", **{"Covers": 2320, "AOV AED": 46.0, "Delivery Mix %": 0.49, "Food Cost %": 0.33, "Labor Cost %": 0.24, "Prime Cost %": 0.57}),
                row("KPI-004", "Apr", "Store A", "Repeat rate push", "Marketing Lead", "CRM", 0.35, 0.29, 164000, 69000, "In Progress", "High", "CRM + POS", **{"Covers": 3980, "AOV AED": 41.2, "Delivery Mix %": 0.30, "Food Cost %": 0.32, "Labor Cost %": 0.29, "Prime Cost %": 0.61}),
                row("KPI-005", "May", "Store D", "Labor pressure", "Finance Lead", "All", 118000, 109500, 109500, 62800, "Blocked", "Critical", "Payroll + POS", **{"Covers": 2550, "AOV AED": 42.9, "Delivery Mix %": 0.27, "Food Cost %": 0.35, "Labor Cost %": 0.36, "Prime Cost %": 0.71}),
                row("KPI-006", "Jun", "Store E", "New menu impact", "Operations Lead", "All", 152000, 160300, 160300, 62000, "Ready", "Medium", "Weekly flash", **{"Covers": 3610, "AOV AED": 44.4, "Delivery Mix %": 0.31, "Food Cost %": 0.30, "Labor Cost %": 0.26, "Prime Cost %": 0.56}),
            ],
        ),
        WorkbookSpec(
            filename="restaurant-promotion-tracker.xlsx",
            folder="growth",
            title="Restaurant Promotion Governance & Performance Tracker",
            subtitle="Offer mechanics, promo cost, redemption, incremental revenue, margin impact, cannibalization, and post-promotion review.",
            category="Promotion governance",
            audience="Restaurant marketers, founders, finance reviewers, and GMs.",
            management_question="Did the promotion create incremental profitable demand, or did it discount revenue that would have happened anyway?",
            dimension_label="Offer Type",
            item_label="Promotion / Offer",
            channel_label="Promotion Channel",
            plan_label="Target Redemptions",
            actual_label="Actual Redemptions",
            value_label="Incremental Revenue AED",
            cost_label="Discount + Media Cost AED",
            efficiency_label="Promo ROI",
            chart_title="Promotion redemptions and incremental revenue",
            gross_margin=0.58,
            min_roi=0.18,
            target_score=0.82,
            assumptions=[
                ("Cannibalization watch band", 0.25, "%", "Promotion governance", "Flags non-incremental revenue", "Finance Lead"),
                ("Minimum contribution ROI", 0.18, "%", "Finance guardrail", "Stops margin-destroying offers", "Finance Lead"),
                ("Promo review deadline", 7, "days after close", "Marketing rule", "Captures learnings", "Marketing Lead"),
                ("Max discount depth", 0.25, "%", "Brand guardrail", "Protects pricing power", "Founder"),
            ],
            definitions=[
                ("Redemption rate", "Actual redemptions divided by eligible audience or target redemptions.", "=Actual Redemptions / Target Redemptions"),
                ("Incremental revenue", "Revenue judged to be caused by the promotion after baseline adjustment.", "Promo Revenue - Baseline Revenue"),
                ("Cannibalization", "Share of promo sales likely taken from full-price sales.", "Cannibalized Revenue / Promo Revenue"),
                ("Promo ROI", "Incremental gross profit less discount and media cost divided by promo cost.", "=(Incremental Revenue x Gross Margin - Cost) / Cost"),
            ],
            actions=[
                ("Discount leakage", "Require margin approval for any offer above the max discount depth.", "Finance Lead", "Critical"),
                ("Weak incrementality", "Use control weeks or comparable stores before declaring the promo successful.", "Marketing Lead", "High"),
                ("Poor post-review", "Add post-promo review within 7 days with keep/cut/change decision.", "Marketing Lead", "Medium"),
            ],
            extra_headers=["Discount %", "Baseline Revenue AED", "Promo Revenue AED", "Cannibalization %", "Post Review Done"],
            rows=[
                row("PROMO-001", "Jan", "Bundle", "Breakfast combo", "Marketing Lead", "In-store", 750, 810, 28400, 8200, "Complete", "High", "POS promo code", **{"Discount %": 0.15, "Baseline Revenue AED": 42000, "Promo Revenue AED": 70400, "Cannibalization %": 0.14, "Post Review Done": "Yes"}),
                row("PROMO-002", "Feb", "BOGO", "Delivery BOGO drink", "Operations Lead", "Aggregator", 620, 690, 21200, 14800, "At Risk", "High", "Aggregator report", **{"Discount %": 0.50, "Baseline Revenue AED": 53000, "Promo Revenue AED": 74200, "Cannibalization %": 0.36, "Post Review Done": "No"}),
                row("PROMO-003", "Mar", "CRM", "Win-back pastry add-on", "Marketing Lead", "WhatsApp", 310, 265, 11600, 2400, "Live", "Medium", "CRM export", **{"Discount %": 0.12, "Baseline Revenue AED": 18000, "Promo Revenue AED": 29600, "Cannibalization %": 0.08, "Post Review Done": "No"}),
                row("PROMO-004", "Apr", "Limited Time", "Pistachio launch", "Founder", "Instagram", 520, 480, 24800, 7600, "Complete", "Medium", "POS + social", **{"Discount %": 0.10, "Baseline Revenue AED": 23000, "Promo Revenue AED": 47800, "Cannibalization %": 0.12, "Post Review Done": "Yes"}),
                row("PROMO-005", "May", "Lunch", "Office lunch voucher", "Store Manager", "Flyers", 400, 210, 7600, 3900, "Blocked", "High", "Voucher log", **{"Discount %": 0.20, "Baseline Revenue AED": 21000, "Promo Revenue AED": 28600, "Cannibalization %": 0.22, "Post Review Done": "No"}),
                row("PROMO-006", "Jun", "Family", "Weekend family box", "Marketing Lead", "Meta", 280, 340, 19800, 5400, "Ready", "Medium", "Forecast", **{"Discount %": 0.18, "Baseline Revenue AED": 26000, "Promo Revenue AED": 45800, "Cannibalization %": 0.11, "Post Review Done": "No"}),
            ],
        ),
        WorkbookSpec(
            filename="store-launch-marketing-checklist.xlsx",
            folder="growth",
            title="Store Launch Marketing Control Tower",
            subtitle="Pre-launch, launch week, and post-launch tasks with owners, deadlines, dependencies, approvals, readiness score, and risk log.",
            category="Store launch",
            audience="Expansion teams, founders, store managers, franchisees, and launch marketers.",
            management_question="Is the store truly ready to open with demand, local visibility, approvals, and follow-through?",
            dimension_label="Launch Phase",
            item_label="Launch Workstream",
            channel_label="Owner Function",
            plan_label="Target Readiness %",
            actual_label="Current Readiness %",
            value_label="Expected First-Month Sales AED",
            cost_label="Launch Cost AED",
            efficiency_label="Launch Payoff",
            chart_title="Launch readiness by phase",
            gross_margin=0.60,
            min_roi=0.20,
            target_score=0.88,
            assumptions=[
                ("Pre-launch window", 60, "days", "Launch playbook", "Starts visibility early", "Marketing Lead"),
                ("Soft opening review", 7, "days before open", "Ops rule", "Catches service issues", "Operations Lead"),
                ("Aggregator setup deadline", 14, "days before open", "Delivery setup", "Protects off-premise launch", "Operations Lead"),
                ("Post-launch review window", 30, "days", "Management cadence", "Stops momentum decay", "Founder"),
            ],
            definitions=[
                ("Launch readiness", "Weighted completion of marketing, operations, permits, signage, CRM, aggregator, and PR tasks.", "Average readiness inputs"),
                ("Dependency", "A blocker that must be completed before the task can move.", "Task relationship"),
                ("Launch risk", "Issue that can delay opening, weaken demand, or damage first impression.", "Risk log status"),
                ("Control tower", "Single view of owner, deadline, status, risk, and next action.", "Governance table"),
            ],
            actions=[
                ("Approval blocker", "Escalate landlord, permit, signage, or mall approval delays weekly.", "Operations Lead", "Critical"),
                ("Weak pre-launch buzz", "Start Google, local PR, influencer, and community outreach 6-8 weeks before opening.", "Marketing Lead", "High"),
                ("No post-launch plan", "Lock first 30-day retention and review cadence before opening week.", "Founder", "High"),
            ],
            extra_headers=["Deadline", "Dependency", "Permit / Approval", "Launch Week Critical", "Risk Level"],
            rows=[
                row("LAUNCH-001", "Jan", "Pre-launch", "Google profile setup", "Marketing Lead", "Local SEO", 0.95, 0.80, 38000, 1800, "In Progress", "High", "GBP draft", **{"Deadline": date(2026, 1, 20), "Dependency": "Lease docs", "Permit / Approval": "No", "Launch Week Critical": "Yes", "Risk Level": "High"}),
                row("LAUNCH-002", "Jan", "Pre-launch", "Exterior signage and mall approvals", "Operations Lead", "Operations", 0.90, 0.55, 52000, 12000, "Blocked", "Critical", "Landlord email", **{"Deadline": date(2026, 1, 25), "Dependency": "Landlord NOC", "Permit / Approval": "Pending", "Launch Week Critical": "Yes", "Risk Level": "Critical"}),
                row("LAUNCH-003", "Feb", "Launch week", "Soft opening invite list", "Founder", "PR", 0.85, 0.72, 24000, 3500, "At Risk", "High", "Invite sheet", **{"Deadline": date(2026, 2, 5), "Dependency": "Menu final", "Permit / Approval": "Yes", "Launch Week Critical": "Yes", "Risk Level": "High"}),
                row("LAUNCH-004", "Feb", "Launch week", "Influencer tasting", "Agency", "Influencer", 0.80, 0.76, 31000, 7800, "Ready", "Medium", "Creator list", **{"Deadline": date(2026, 2, 8), "Dependency": "Food photography", "Permit / Approval": "Yes", "Launch Week Critical": "No", "Risk Level": "Medium"}),
                row("LAUNCH-005", "Mar", "Post-launch", "CRM first-visit follow-up", "Marketing Lead", "WhatsApp", 0.90, 0.66, 27000, 2100, "In Progress", "High", "Automation setup", **{"Deadline": date(2026, 3, 8), "Dependency": "POS customer capture", "Permit / Approval": "Yes", "Launch Week Critical": "No", "Risk Level": "High"}),
                row("LAUNCH-006", "Mar", "Post-launch", "30-day performance review", "Finance Lead", "Management", 1.00, 0.25, 0, 0, "Not Started", "Medium", "Review template", **{"Deadline": date(2026, 3, 30), "Dependency": "Weekly sales", "Permit / Approval": "Yes", "Launch Week Critical": "No", "Risk Level": "Medium"}),
            ],
        ),
        WorkbookSpec(
            filename="weekly-flash-report-template.xlsx",
            folder="finance",
            title="CEO / Investor Weekly Flash Report",
            subtitle="Weekly trading summary, KPI variance, risks, opportunities, cash/sales indicators, store performance, marketing impact, and action tracker.",
            category="Weekly flash",
            audience="CEOs, founders, finance leads, investors, and operators reviewing weekly performance.",
            management_question="What happened this week, why did it happen, what is at risk, and what action is management taking?",
            dimension_label="Store / Function",
            item_label="Weekly KPI / Issue",
            channel_label="Revenue Channel",
            plan_label="Weekly Target AED / KPI",
            actual_label="Weekly Actual AED / KPI",
            value_label="Net Sales / Impact AED",
            cost_label="Cost / Leakage AED",
            efficiency_label="Weekly Health",
            chart_title="Weekly target vs actual trend",
            gross_margin=0.62,
            min_roi=0.18,
            target_score=0.85,
            assumptions=[
                ("Weekly sales materiality", 0.08, "% variance", "Management rule", "Requires commentary", "Finance Lead"),
                ("Prime cost warning", 0.65, "% of sales", "Restaurant finance rule", "Flags margin risk", "Finance Lead"),
                ("Cash watch threshold", 2, "weeks cover", "Treasury guardrail", "Flags funding pressure", "Founder"),
                ("Action owner SLA", 7, "days", "Management cadence", "Keeps review actionable", "Operations Lead"),
            ],
            definitions=[
                ("Weekly variance", "Actual weekly value compared with target.", "Actual - Target"),
                ("Sales vs target", "How trading performed against weekly target.", "=Actual Sales / Target Sales - 1"),
                ("Operational red flag", "Issue requiring management action before month-end.", "Risk threshold"),
                ("Flash report", "Short weekly view of performance, risk, opportunity, and next actions.", "Management pack"),
            ],
            actions=[
                ("Sales miss", "Separate traffic, AOV, delivery mix, and discount causes before deciding action.", "Finance Lead", "High"),
                ("Cost pressure", "Review labor scheduling, food cost, waste, and promos immediately.", "Operations Lead", "Critical"),
                ("Unclosed actions", "Carry forward every owner/action until closed with evidence.", "Founder", "High"),
            ],
            extra_headers=["Week Start", "Covers", "AOV AED", "Food Cost %", "Labor Cost %", "EBITDA %"],
            rows=[
                row("FLASH-001", "Jan", "Store A", "Net sales week 1", "Store Manager", "All", 46000, 49200, 49200, 19800, "Complete", "High", "POS export", **{"Week Start": date(2026, 1, 5), "Covers": 1180, "AOV AED": 41.7, "Food Cost %": 0.31, "Labor Cost %": 0.26, "EBITDA %": 0.15}),
                row("FLASH-002", "Jan", "Store B", "Labor overrun", "Operations Lead", "Dine-in", 39000, 35500, 35500, 18100, "At Risk", "High", "Payroll schedule", **{"Week Start": date(2026, 1, 5), "Covers": 820, "AOV AED": 43.3, "Food Cost %": 0.33, "Labor Cost %": 0.34, "EBITDA %": 0.06}),
                row("FLASH-003", "Feb", "Delivery", "Aggregator margin", "Finance Lead", "Aggregator", 22000, 24800, 24800, 11600, "Live", "Medium", "Payout report", **{"Week Start": date(2026, 2, 2), "Covers": 540, "AOV AED": 45.9, "Food Cost %": 0.34, "Labor Cost %": 0.22, "EBITDA %": 0.08}),
                row("FLASH-004", "Mar", "Marketing", "Campaign sales lift", "Marketing Lead", "Meta", 18000, 22500, 22500, 6200, "Complete", "Medium", "Campaign report", **{"Week Start": date(2026, 3, 2), "Covers": 390, "AOV AED": 57.7, "Food Cost %": 0.28, "Labor Cost %": 0.18, "EBITDA %": 0.19}),
                row("FLASH-005", "Apr", "Store C", "Guest count decline", "Store Manager", "All", 42000, 36800, 36800, 15800, "Blocked", "Critical", "Weekly flash notes", **{"Week Start": date(2026, 4, 6), "Covers": 760, "AOV AED": 48.4, "Food Cost %": 0.32, "Labor Cost %": 0.31, "EBITDA %": 0.04}),
                row("FLASH-006", "May", "Cash", "Supplier payment pressure", "Finance Lead", "Management", 0, 0, 0, 28000, "At Risk", "Critical", "AP aging", **{"Week Start": date(2026, 5, 4), "Covers": 0, "AOV AED": 0, "Food Cost %": 0, "Labor Cost %": 0, "EBITDA %": 0}),
            ],
        ),
    ]


def extend_specs(specs: list[WorkbookSpec]) -> list[WorkbookSpec]:
    """Add the remaining specialized growth workbooks."""
    specs.extend(
        [
            WorkbookSpec(
                filename="store-launch-marketing-checklist.xlsx",
                folder="growth",
                title="Store Launch Marketing Control Tower",
                subtitle="Pre-launch, launch week, and post-launch launch readiness workbook.",
                category="Launch control",
                audience="Founders and expansion teams opening a new location.",
                management_question="Is the launch ready across demand, approvals, CRM, aggregator, signage, and review cadence?",
                dimension_label="Launch Phase",
                item_label="Launch Task",
                channel_label="Function",
                plan_label="Target Readiness %",
                actual_label="Current Readiness %",
                value_label="Expected Launch Sales AED",
                cost_label="Launch Cost AED",
                efficiency_label="Launch ROI",
                chart_title="Launch readiness trend",
                gross_margin=0.60,
                min_roi=0.20,
                target_score=0.88,
                assumptions=[
                    ("Pre-launch planning window", 60, "days", "Launch standard", "Protects momentum", "Marketing Lead"),
                    ("Minimum launch readiness", 0.88, "%", "Go-live gate", "Blocks risky opening", "Founder"),
                    ("Aggregator setup deadline", 14, "days", "Ops rule", "Avoids delivery delay", "Operations Lead"),
                    ("Post-launch review window", 30, "days", "Management cadence", "Captures first-month learnings", "Finance Lead"),
                ],
                definitions=[
                    ("Launch readiness", "Weighted completion of launch-critical workstreams.", "Average readiness score"),
                    ("Critical path", "Tasks that can delay or weaken opening if incomplete.", "Dependency logic"),
                    ("Post-launch review", "30-day review of sales, traffic, reviews, marketing, and issues.", "Management meeting pack"),
                    ("Risk log", "Open launch blockers with owners and due dates.", "Issue tracker"),
                ],
                actions=[
                    ("Approval delay", "Escalate landlord, permit, signage, and mall approvals weekly.", "Operations Lead", "Critical"),
                    ("No launch demand", "Start GBP, local outreach, PR, and influencer work before the store opens.", "Marketing Lead", "High"),
                    ("Weak follow-through", "Lock the first 30-day CRM and review plan before opening.", "Founder", "High"),
                ],
                extra_headers=["Deadline", "Dependency", "Approval Status", "Launch Critical", "Risk Level"],
                rows=[
                    row("SLC-001", "Jan", "Pre-launch", "Google profile and local listings", "Marketing Lead", "Local SEO", 0.95, 0.82, 42000, 1800, "In Progress", "High", "GBP draft", **{"Deadline": date(2026, 1, 20), "Dependency": "Lease docs", "Approval Status": "Pending", "Launch Critical": "Yes", "Risk Level": "High"}),
                    row("SLC-002", "Jan", "Pre-launch", "Exterior signage approval", "Operations Lead", "Operations", 0.90, 0.55, 52000, 12000, "Blocked", "Critical", "Landlord thread", **{"Deadline": date(2026, 1, 25), "Dependency": "Landlord NOC", "Approval Status": "Blocked", "Launch Critical": "Yes", "Risk Level": "Critical"}),
                    row("SLC-003", "Feb", "Launch week", "Soft opening invite list", "Founder", "PR", 0.85, 0.72, 24000, 3500, "At Risk", "High", "Invite sheet", **{"Deadline": date(2026, 2, 5), "Dependency": "Menu final", "Approval Status": "In Review", "Launch Critical": "Yes", "Risk Level": "High"}),
                    row("SLC-004", "Feb", "Launch week", "Influencer tasting", "Agency", "Influencer", 0.80, 0.76, 31000, 7800, "Ready", "Medium", "Creator list", **{"Deadline": date(2026, 2, 8), "Dependency": "Food photography", "Approval Status": "Approved", "Launch Critical": "No", "Risk Level": "Medium"}),
                    row("SLC-005", "Mar", "Post-launch", "CRM first-visit follow-up", "Marketing Lead", "WhatsApp", 0.90, 0.66, 27000, 2100, "In Progress", "High", "Automation setup", **{"Deadline": date(2026, 3, 8), "Dependency": "POS customer capture", "Approval Status": "Pending", "Launch Critical": "No", "Risk Level": "High"}),
                ],
            ),
            WorkbookSpec(
                filename="weekly-flash-report-template.xlsx",
                folder="finance",
                title="CEO / Investor Weekly Flash Report",
                subtitle="Weekly performance summary with KPI variance, risks, opportunities, marketing impact, and action ownership.",
                category="Weekly management report",
                audience="CEOs, investors, founders, finance leads, and operators.",
                management_question="What changed this week, what is at risk, and what must management do next?",
                dimension_label="Store / Function",
                item_label="Weekly KPI / Issue",
                channel_label="Trading Channel",
                plan_label="Weekly Target AED / KPI",
                actual_label="Weekly Actual AED / KPI",
                value_label="Net Sales / Impact AED",
                cost_label="Cost / Leakage AED",
                efficiency_label="Weekly Health",
                chart_title="Weekly target vs actual trend",
                gross_margin=0.62,
                min_roi=0.18,
                target_score=0.85,
                assumptions=[
                    ("Weekly sales materiality", 0.08, "% variance", "Management rule", "Requires commentary", "Finance Lead"),
                    ("Prime cost warning", 0.65, "% of sales", "Restaurant finance rule", "Flags margin risk", "Finance Lead"),
                    ("Cash watch threshold", 2, "weeks cover", "Treasury guardrail", "Flags funding pressure", "Founder"),
                    ("Action owner SLA", 7, "days", "Management cadence", "Keeps review actionable", "Operations Lead"),
                ],
                definitions=[
                    ("Weekly variance", "Actual weekly value compared with target.", "Actual - Target"),
                    ("Sales vs target", "How trading performed against weekly target.", "=Actual Sales / Target Sales - 1"),
                    ("Operational red flag", "Issue requiring management action before month-end.", "Risk threshold"),
                    ("Flash report", "Short weekly view of performance, risk, opportunity, and next actions.", "Management pack"),
                ],
                actions=[
                    ("Sales miss", "Separate traffic, AOV, delivery mix, and discount causes before deciding action.", "Finance Lead", "High"),
                    ("Cost pressure", "Review labor scheduling, food cost, waste, and promos immediately.", "Operations Lead", "Critical"),
                    ("Unclosed actions", "Carry forward every owner/action until closed with evidence.", "Founder", "High"),
                ],
                extra_headers=["Week Start", "Covers", "AOV AED", "Food Cost %", "Labor Cost %", "EBITDA %"],
                rows=[
                    row("FLASH-001", "Jan", "Store A", "Net sales week 1", "Store Manager", "All", 46000, 49200, 49200, 19800, "Complete", "High", "POS export", **{"Week Start": date(2026, 1, 5), "Covers": 1180, "AOV AED": 41.7, "Food Cost %": 0.31, "Labor Cost %": 0.26, "EBITDA %": 0.15}),
                    row("FLASH-002", "Jan", "Store B", "Labor overrun", "Operations Lead", "Dine-in", 39000, 35500, 35500, 18100, "At Risk", "High", "Payroll schedule", **{"Week Start": date(2026, 1, 5), "Covers": 820, "AOV AED": 43.3, "Food Cost %": 0.33, "Labor Cost %": 0.34, "EBITDA %": 0.06}),
                    row("FLASH-003", "Feb", "Delivery", "Aggregator margin", "Finance Lead", "Aggregator", 22000, 24800, 24800, 11600, "Live", "Medium", "Payout report", **{"Week Start": date(2026, 2, 2), "Covers": 540, "AOV AED": 45.9, "Food Cost %": 0.34, "Labor Cost %": 0.22, "EBITDA %": 0.08}),
                    row("FLASH-004", "Mar", "Marketing", "Campaign sales lift", "Marketing Lead", "Meta", 18000, 22500, 22500, 6200, "Complete", "Medium", "Campaign report", **{"Week Start": date(2026, 3, 2), "Covers": 390, "AOV AED": 57.7, "Food Cost %": 0.28, "Labor Cost %": 0.18, "EBITDA %": 0.19}),
                    row("FLASH-005", "Apr", "Store C", "Guest count decline", "Store Manager", "All", 42000, 36800, 36800, 15800, "Blocked", "Critical", "Weekly flash notes", **{"Week Start": date(2026, 4, 6), "Covers": 760, "AOV AED": 48.4, "Food Cost %": 0.32, "Labor Cost %": 0.31, "EBITDA %": 0.04}),
                ],
            ),
        ]
    )
    deduped: dict[str, WorkbookSpec] = {}
    for spec in specs:
        deduped[spec.filename] = spec
    return list(deduped.values())


def clone_spec(template: WorkbookSpec, filename: str, title: str, subtitle: str, category: str, dimension: str, item: str, rows: list[SampleRow], extra_headers: list[str], definitions: list[tuple[str, str, str]], actions: list[tuple[str, str, str, str]], assumptions: list[tuple[str, Any, str, str, str, str]], chart_title: str, folder: str = "growth") -> WorkbookSpec:
    return WorkbookSpec(
        filename=filename,
        folder=folder,
        title=title,
        subtitle=subtitle,
        category=category,
        audience=template.audience,
        management_question=template.management_question,
        dimension_label=dimension,
        item_label=item,
        channel_label=template.channel_label,
        plan_label=template.plan_label,
        actual_label=template.actual_label,
        value_label=template.value_label,
        cost_label=template.cost_label,
        efficiency_label=template.efficiency_label,
        chart_title=chart_title,
        gross_margin=template.gross_margin,
        min_roi=template.min_roi,
        target_score=template.target_score,
        assumptions=assumptions,
        definitions=definitions,
        actions=actions,
        extra_headers=extra_headers,
        rows=rows,
    )


def add_missing_specs(specs: list[WorkbookSpec]) -> list[WorkbookSpec]:
    """Add the few workbooks whose structures differ mainly by business lens."""
    base = specs[0]
    specs.extend(
        [
            WorkbookSpec(
                filename="local-store-marketing-planner.xlsx",
                folder="growth",
                title="Local Store Marketing Strategy Planner",
                subtitle="Trade-area mapping, local segments, partnerships, hyperlocal campaigns, footfall estimates, budget, and ROI tracking.",
                category="Local store marketing",
                audience="Store managers, local marketers, franchisees, and single-outlet founders.",
                management_question="Which local moves will bring realistic footfall and repeat visits for this exact trade area?",
                dimension_label="Trade Area Segment",
                item_label="Local Activity / Partner",
                channel_label="Activation Channel",
                plan_label="Target Footfall / Leads",
                actual_label="Actual Footfall / Leads",
                value_label="Estimated Sales Impact AED",
                cost_label="Local Marketing Cost AED",
                efficiency_label="Local ROI",
                chart_title="Local footfall plan vs actual",
                gross_margin=0.61,
                min_roi=0.20,
                target_score=0.82,
                assumptions=[
                    ("Trade area radius", 3, "km", "Local planning rule", "Defines target clusters", "Store Manager"),
                    ("Conversion from sample to visit", 0.18, "%", "Operator assumption", "Estimates local ROI", "Marketing Lead"),
                    ("Office partnership minimum", 5, "partners", "Local playbook", "Builds weekday demand", "Store Manager"),
                    ("Local campaign review cadence", 14, "days", "Management rhythm", "Stops dead activities", "Marketing Lead"),
                ],
                definitions=[
                    ("Trade area", "The practical local catchment where customers can visit or order from the store.", "Radius + demand clusters"),
                    ("Footfall estimate", "Expected store visits generated by a local activity.", "Reach x conversion rate"),
                    ("Partnership value", "Sales impact from schools, offices, residential communities, gyms, malls, or local businesses.", "Estimated orders x AOV"),
                    ("Local ROI", "Gross profit from local sales impact less activity cost, divided by cost.", "=(Sales impact x Gross Margin - Cost) / Cost"),
                ],
                actions=[
                    ("Weak office pipeline", "Build a target list of nearby offices with decision-maker, offer, and follow-up date.", "Store Manager", "High"),
                    ("Low conversion", "Change offer from generic discount to specific trial reason by segment.", "Marketing Lead", "Medium"),
                    ("Untracked activities", "Require QR, coupon code, or POS tag for every local activation.", "Finance Lead", "High"),
                ],
                extra_headers=["Cluster Type", "Distance Km", "Partner Count", "Conversion Rate %", "AOV AED"],
                rows=[
                    row("LSM-001", "Jan", "Office cluster", "Lunch tasting at business tower", "Store Manager", "In-store", 220, 185, 13800, 3100, "Complete", "High", "Sampling log", **{"Cluster Type": "Office", "Distance Km": 0.7, "Partner Count": 3, "Conversion Rate %": 0.16, "AOV AED": 42}),
                    row("LSM-002", "Feb", "Residential", "Weekend family bundle flyer", "Marketing Lead", "Flyers", 160, 118, 7200, 1800, "Live", "Medium", "Coupon code", **{"Cluster Type": "Residential", "Distance Km": 1.5, "Partner Count": 1, "Conversion Rate %": 0.09, "AOV AED": 61}),
                    row("LSM-003", "Mar", "School", "After-school snack tie-up", "Store Manager", "OOH", 190, 142, 8600, 2600, "In Progress", "High", "School approvals", **{"Cluster Type": "School", "Distance Km": 1.1, "Partner Count": 2, "Conversion Rate %": 0.12, "AOV AED": 37}),
                    row("LSM-004", "Apr", "Gym", "Protein coffee collaboration", "Founder", "PR", 100, 125, 11200, 2400, "Complete", "Medium", "Partner report", **{"Cluster Type": "Fitness", "Distance Km": 0.5, "Partner Count": 1, "Conversion Rate %": 0.22, "AOV AED": 48}),
                    row("LSM-005", "May", "Mall traffic", "Entrance sampling", "Operations Lead", "In-store", 300, 165, 9900, 4200, "At Risk", "High", "Mall permit issue", **{"Cluster Type": "Mall", "Distance Km": 0.0, "Partner Count": 1, "Conversion Rate %": 0.07, "AOV AED": 45}),
                ],
            ),
        ]
    )
    order = [
        "annual-marketing-budget-planner.xlsx",
        "annual-restaurant-marketing-plan.xlsx",
        "cafe-content-calendar.xlsx",
        "crm-loyalty-campaign-planner.xlsx",
        "delivery-aggregator-audit.xlsx",
        "franchise-campaign-rollout-tracker.xlsx",
        "google-business-profile-checklist.xlsx",
        "local-store-marketing-planner.xlsx",
        "marketing-roi-calculator.xlsx",
        "menu-launch-and-offer-planner.xlsx",
        "restaurant-brand-positioning-brief.xlsx",
        "restaurant-kpi-dashboard-template.xlsx",
        "restaurant-promotion-tracker.xlsx",
        "store-launch-marketing-checklist.xlsx",
        "weekly-flash-report-template.xlsx",
    ]
    by_name = {spec.filename: spec for spec in specs}
    missing = [name for name in order if name not in by_name]
    if missing:
        raise RuntimeError(f"Missing workbook specs: {missing}")
    return [by_name[name] for name in order]


def set_title(ws, title: str, subtitle: str, last_col: int = 12) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.fill = HEADER_FILL
    cell.font = Font(name=FONTS["heading"], size=20, bold=True, color=COLORS["white"])
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = HEADER_FILL
        ws.cell(2, col).fill = HEADER_FILL
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=last_col)
    sub = ws.cell(3, 1, subtitle)
    sub.fill = SECTION_FILL
    sub.font = Font(name=FONTS["base"], size=10, color=COLORS["muted"])
    sub.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, last_col + 1):
        ws.cell(3, col).fill = SECTION_FILL
        ws.cell(4, col).fill = SECTION_FILL
    for row_idx in range(1, 5):
        for col in range(1, last_col + 1):
            ws.cell(row_idx, col).border = BORDER
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24


def style_range(ws, min_row: int, min_col: int, max_row: int, max_col: int, fill: PatternFill | None = None, font: Font | None = None, alignment: Alignment | None = None) -> None:
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment


def add_section(ws, row_idx: int, title: str, last_col: int = 8) -> None:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
    cell = ws.cell(row_idx, 1, title)
    cell.fill = SECTION_FILL
    cell.font = Font(name=FONTS["heading"], size=12, bold=True, color=COLORS["charcoal"])
    cell.alignment = Alignment(vertical="center")
    for col in range(1, last_col + 1):
        ws.cell(row_idx, col).border = Border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN)


def set_widths(ws, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name[:250], ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def list_validation(values: list[str]) -> DataValidation:
    formula = '"' + ",".join(values) + '"'
    return DataValidation(type="list", formula1=formula, allow_blank=True)


def number_format_for(header: str) -> str:
    h = header.lower()
    if "date" in h or "deadline" in h or "week start" in h:
        return "dd-mmm-yyyy"
    if "aed" in h or "revenue" in h or "cost" in h or "budget" in h or "spend" in h or "sales" in h or "value" in h or "impact" in h or "cac" in h or "aov" in h or "ltv" in h:
        return '"AED" #,##0;[Red]-"AED" #,##0;-'
    if "%" in header or "rate" in h or "margin" in h or "mix" in h or "score" in h or "readiness" in h or "roi" in h or "health" in h or "lift" in h:
        return "0.0%;[Red](0.0%);-"
    if "roas" in h:
        return "0.0x"
    if any(word in h for word in ["units", "count", "customers", "clicks", "impressions", "covers", "days", "hrs", "hours", "min", "rank", "rating", "posts", "redemptions"]):
        return "#,##0.0" if "rating" in h else "#,##0"
    return "General"


def create_readme(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("README")
    set_title(ws, spec.title, "How to use this workbook and what each sheet is for.", 10)
    set_widths(ws, [22, 32, 48, 22, 22, 22, 22, 22, 22, 22])
    add_section(ws, 6, "Workbook purpose", 10)
    ws["A7"] = spec.management_question
    ws.merge_cells("A7:J8")
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A7"].font = Font(name=FONTS["base"], size=11, color=COLORS["charcoal"])
    style_range(ws, 7, 1, 8, 10, PatternFill("solid", fgColor=COLORS["panel"]))

    add_section(ws, 10, "How to use", 10)
    steps = [
        "1. Read README and Document_Control first so the purpose, owner, and version are clear.",
        "2. Update Assumptions before changing the operating data. Blue input cells are meant to be edited.",
        "3. Replace sample rows in Inputs with your own store, campaign, customer, or finance data.",
        "4. Use Calc and Dashboard for management review. Do not overwrite formula cells unless you are redesigning the model.",
        "5. Review Checks before sharing the workbook with investors, auditors, franchisees, or senior management.",
        "6. Use Scenarios and Action_Plan to convert the numbers into a management decision.",
    ]
    for offset, step in enumerate(steps, start=11):
        ws.cell(offset, 1, step)
        ws.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=10)
        style_range(ws, offset, 1, offset, 10, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))

    add_section(ws, 19, "Workbook map", 10)
    rows = [
        ("Document_Control", "Version, owner, review status, and change log."),
        ("Assumptions", "Editable model assumptions, thresholds, and source notes."),
        ("Inputs", "Primary user-editable operating table with sample rows."),
        ("Calc", "Formula-led monthly summary and decision signals."),
        ("Dashboard", "Board-style summary with KPI cards and chart."),
        ("Checks", "Audit checks for missing inputs, overspend, low ROI, duplicate records, and open risks."),
        ("Definitions", "KPI glossary and source library."),
        ("Scenarios", "Base, Conservative, and Aggressive case view."),
        ("Action_Plan", "Management actions, owners, priorities, and due dates."),
    ]
    ws.append(["Sheet", "Purpose"])
    for idx, item in enumerate(rows, start=21):
        ws.cell(idx, 1, item[0])
        ws.cell(idx, 2, item[1])
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=10)
    style_range(ws, 20, 1, 20, 10, HEADER_FILL, Font(name=FONTS["base"], size=10, bold=True, color=COLORS["white"]))
    style_range(ws, 21, 1, 29, 10, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))

    add_section(ws, 31, "Color legend", 10)
    legend = [
        ("Blue fill / blue font", "User input or assumption cell."),
        ("White / black font", "Formula output or calculated result."),
        ("Green fill", "Internal workbook link or dashboard/status reference."),
        ("Amber / red conditional fill", "Review required, variance, risk, or failed check."),
    ]
    for idx, (label, desc) in enumerate(legend, start=32):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, desc)
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=10)
    style_range(ws, 32, 1, 35, 10, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    ws["A32"].fill = INPUT_FILL
    ws["A33"].fill = FORMULA_FILL
    ws["A34"].fill = LINK_FILL
    ws["A35"].fill = YELLOW_FILL


def create_document_control(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Document_Control")
    set_title(ws, "Document Control", "Ownership, version control, review status, and change log.", 8)
    set_widths(ws, [24, 34, 24, 28, 28, 28, 24, 36])
    add_section(ws, 6, "Control summary", 8)
    rows = [
        ("Workbook", spec.title, "Filename", spec.filename),
        ("Version", VERSION, "Prepared by", "Ashmo.io"),
        ("Template date", TEMPLATE_DATE, "Review status", "Issued for user customization"),
        ("Model owner", "User to assign", "Audience", spec.audience),
        ("Primary question", spec.management_question, "Currency", "AED unless changed in Assumptions"),
    ]
    for r, values in enumerate(rows, start=7):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)
    style_range(ws, 7, 1, 11, 8, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    for r in range(7, 12):
        for c in (1, 3, 5, 7):
            ws.cell(r, c).font = Font(name=FONTS["base"], bold=True, color=COLORS["muted"])
        for c in (2, 4, 6, 8):
            ws.cell(r, c).fill = INPUT_FILL
            ws.cell(r, c).font = Font(name=FONTS["base"], color="0000FF")

    add_section(ws, 14, "Change log", 8)
    headers = ["Date", "Version", "Changed by", "Change", "Reason", "Reviewer", "Status", "Notes"]
    for c, header in enumerate(headers, start=1):
        ws.cell(15, c, header)
    changes = [
        (TEMPLATE_DATE, VERSION, "Ashmo.io", "Premium rebuild", "Investor/audit-ready framework upgrade", "User", "Open", "Replace sample data before live use."),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
    ]
    for r, values in enumerate(changes, start=16):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)
    style_range(ws, 15, 1, 15, 8, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 16, 1, 18, 8, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    add_table(ws, f"tbl_{spec.slug}_changelog", 15, 1, 18, 8)


def create_assumptions(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Assumptions")
    set_title(ws, "Assumptions", "Editable assumptions, review thresholds, and source notes.", 8)
    set_widths(ws, [34, 16, 16, 42, 34, 20, 18, 18])
    add_section(ws, 6, "Core assumptions and thresholds", 8)
    headers = ["Assumption", "Value", "Unit", "Source / Basis", "Impact on model", "Owner", "Last reviewed", "Notes"]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    core = [
        ("Currency", "AED", "text", "Default workbook convention", "All financial values use AED unless changed", "Finance Lead"),
        ("Reporting period", "Monthly", "text", "Workbook standard", "Calc and dashboard summarize by month", "Finance Lead"),
        ("Gross margin / contribution margin %", spec.gross_margin, "%", "User to replace with actual margin", "Used in ROI and contribution formulas", "Finance Lead"),
        ("Material variance threshold", 0.10, "%", "Management review policy", "Flags overspend or underperformance", "Finance Lead"),
        ("Minimum ROI / health threshold", spec.min_roi, "%", "Management hurdle", "Flags low-return activity", "Finance Lead"),
        ("Target readiness / compliance score", spec.target_score, "%", "Operator target", "Flags readiness and quality gaps", "Operations Lead"),
        ("Default review cadence", "Monthly", "text", "Management rhythm", "Sets operating review timing", "Founder"),
    ]
    full = core + spec.assumptions
    for r, values in enumerate(full, start=8):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)
        ws.cell(r, 7, TEMPLATE_DATE)
    style_range(ws, 7, 1, 7, 8, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 8, 1, 8 + len(full) - 1, 8, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    for row_idx in range(8, 8 + len(full)):
        ws.cell(row_idx, 2).fill = INPUT_FILL
        ws.cell(row_idx, 2).font = Font(name=FONTS["base"], color="0000FF")
        ws.cell(row_idx, 2).comment = Comment("Editable assumption. Update before live management use.", "Ashmo.io")
        if isinstance(ws.cell(row_idx, 2).value, float):
            ws.cell(row_idx, 2).number_format = "0.0%"
    add_table(ws, f"tbl_{spec.slug}_assumptions", 7, 1, 8 + len(full) - 1, 8)


def input_headers(spec: WorkbookSpec) -> list[str]:
    return [
        "Record ID",
        "Month",
        spec.dimension_label,
        spec.item_label,
        "Owner",
        spec.channel_label,
        spec.plan_label,
        spec.actual_label,
        spec.value_label,
        spec.cost_label,
        *spec.extra_headers,
        "Status",
        "Priority",
        "Notes / Evidence",
        "Variance",
        "Variance %",
        spec.efficiency_label,
        "Risk Flag",
    ]


def header_map(headers: list[str]) -> dict[str, str]:
    return {header: get_column_letter(index) for index, header in enumerate(headers, start=1)}


def create_inputs(wb: Workbook, spec: WorkbookSpec) -> tuple[list[str], dict[str, str]]:
    ws = wb.create_sheet("Inputs")
    headers = input_headers(spec)
    cols = header_map(headers)
    last_col = len(headers)
    set_title(ws, "Inputs", "Replace sample rows with your own data. Blue cells are editable inputs; formula columns should be left intact.", last_col)
    set_widths(ws, [14, 10, 24, 30, 18, 18, 16, 16, 18, 18] + [16] * len(spec.extra_headers) + [15, 12, 34, 16, 14, 16, 14])
    start_row = 7
    data_start = 8
    data_end = 57
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, c, header)
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONTS["base"], bold=True, color=COLORS["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    formula_headers = {"Variance", "Variance %", spec.efficiency_label, "Risk Flag"}
    sample_by_row = {data_start + idx: sample for idx, sample in enumerate(spec.rows)}
    for row_idx in range(data_start, data_end + 1):
        sample = sample_by_row.get(row_idx)
        values: dict[str, Any] = {}
        if sample:
            values.update(
                {
                    "Record ID": sample.record_id,
                    "Month": sample.month,
                    spec.dimension_label: sample.dimension,
                    spec.item_label: sample.item,
                    "Owner": sample.owner,
                    spec.channel_label: sample.channel,
                    spec.plan_label: sample.plan,
                    spec.actual_label: sample.actual,
                    spec.value_label: sample.value,
                    spec.cost_label: sample.cost,
                    "Status": sample.status,
                    "Priority": sample.priority,
                    "Notes / Evidence": sample.evidence,
                }
            )
            values.update(sample.extras)
        for c, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header in formula_headers:
                cell.fill = FORMULA_FILL
                cell.font = Font(name=FONTS["base"], color=COLORS["charcoal"])
                cell.protection = Protection(locked=True)
            else:
                cell.value = values.get(header, "")
                cell.fill = INPUT_FILL
                cell.font = Font(name=FONTS["base"], color="0000FF")
                cell.protection = Protection(locked=False)
            cell.number_format = number_format_for(header)

        plan = f"{cols[spec.plan_label]}{row_idx}"
        actual = f"{cols[spec.actual_label]}{row_idx}"
        value = f"{cols[spec.value_label]}{row_idx}"
        cost = f"{cols[spec.cost_label]}{row_idx}"
        status = f"{cols['Status']}{row_idx}"
        priority = f"{cols['Priority']}{row_idx}"
        variance = f"{cols['Variance']}{row_idx}"
        variance_pct = f"{cols['Variance %']}{row_idx}"
        efficiency = f"{cols[spec.efficiency_label]}{row_idx}"
        ws[variance] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({actual}-{plan},""))'
        ws[variance_pct] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({variance}/ABS({plan}),""))'
        ws[efficiency] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR(({value}*Assumptions!$B$10-{cost})/{cost},""))'
        ws[f"{cols['Risk Flag']}{row_idx}"] = (
            f'=IF({cols["Record ID"]}{row_idx}="","",IF(OR({status}="Blocked",{priority}="Critical",'
            f'{efficiency}<Assumptions!$B$12,ABS({variance_pct})>Assumptions!$B$11),"REVIEW","OK"))'
        )

        # Optional row-level formulas for domain-specific metrics.
        if "CAC AED" in cols and any("Customers" in h for h in headers):
            customer_header = next((h for h in headers if "Customers" in h), None)
            if customer_header:
                ws[f"{cols['CAC AED']}{row_idx}"] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({cost}/{cols[customer_header]}{row_idx},""))'
        if "ROAS x" in cols:
            ws[f"{cols['ROAS x']}{row_idx}"] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({value}/{cost},""))'
        if "Payback Months" in cols:
            ws[f"{cols['Payback Months']}{row_idx}"] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({cost}/(({value}*Assumptions!$B$10)/12),""))'
        if "Gross Margin %" in cols and "Food Cost AED" in cols and "Selling Price AED" in cols:
            ws[f"{cols['Gross Margin %']}{row_idx}"] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR(({cols["Selling Price AED"]}{row_idx}-{cols["Food Cost AED"]}{row_idx})/{cols["Selling Price AED"]}{row_idx},""))'
        if "Breakeven Units" in cols and "Food Cost AED" in cols and "Selling Price AED" in cols:
            ws[f"{cols['Breakeven Units']}{row_idx}"] = f'=IF({cols["Record ID"]}{row_idx}="","",IFERROR({cost}/({cols["Selling Price AED"]}{row_idx}-{cols["Food Cost AED"]}{row_idx}),""))'

    validations = [
        ("Month", MONTH_VALUES),
        ("Status", STATUS_VALUES),
        ("Priority", PRIORITY_VALUES),
        ("Owner", OWNER_VALUES),
        (spec.channel_label, CHANNEL_VALUES),
    ]
    for header, values in validations:
        if header not in cols:
            continue
        dv = list_validation(values)
        ws.add_data_validation(dv)
        dv.add(f"{cols[header]}{data_start}:{cols[header]}{data_end}")
    for header in spec.extra_headers:
        h = header.lower()
        if any(term in h for term in ["approval", "sign-off", "critical", "review done", "one sentence"]):
            dv = list_validation(["Yes", "No", "Pending", "Blocked", "Approved", "Review"])
            ws.add_data_validation(dv)
            dv.add(f"{cols[header]}{data_start}:{cols[header]}{data_end}")

    ws.freeze_panes = "A8"
    add_table(ws, f"tbl_{spec.slug}_inputs", start_row, 1, data_end, last_col)
    ws.conditional_formatting.add(f"{cols['Risk Flag']}{data_start}:{cols['Risk Flag']}{data_end}", FormulaRule(formula=[f'{cols["Risk Flag"]}{data_start}="REVIEW"'], fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.conditional_formatting.add(f"{cols['Risk Flag']}{data_start}:{cols['Risk Flag']}{data_end}", FormulaRule(formula=[f'{cols["Risk Flag"]}{data_start}="OK"'], fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(f"{cols['Variance %']}{data_start}:{cols['Variance %']}{data_end}", CellIsRule(operator="greaterThan", formula=["Assumptions!$B$11"], fill=PatternFill("solid", fgColor="FEF3C7")))
    ws.conditional_formatting.add(f"{cols[spec.efficiency_label]}{data_start}:{cols[spec.efficiency_label]}{data_end}", CellIsRule(operator="lessThan", formula=["Assumptions!$B$12"], fill=PatternFill("solid", fgColor="FEE2E2")))
    return headers, cols


def create_calc(wb: Workbook, spec: WorkbookSpec, cols: dict[str, str]) -> None:
    ws = wb.create_sheet("Calc")
    set_title(ws, "Calc", "Formula-led monthly roll-up. Keep formulas intact for auditability.", 13)
    set_widths(ws, [10, 18, 18, 18, 18, 18, 16, 14, 16, 16, 14, 22, 24])
    headers = [
        "Month",
        "Plan / Target",
        "Actual / Progress",
        "Revenue / Value",
        "Cost / Investment",
        "Gross Profit / Benefit",
        "Variance",
        "Variance %",
        spec.efficiency_label,
        "Completion %",
        "Open Risks",
        "Top Area",
        "Management Signal",
    ]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    style_range(ws, 7, 1, 7, 13, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))

    month_range = f"Inputs!${cols['Month']}$8:${cols['Month']}$57"
    plan_range = f"Inputs!${cols[spec.plan_label]}$8:${cols[spec.plan_label]}$57"
    actual_range = f"Inputs!${cols[spec.actual_label]}$8:${cols[spec.actual_label]}$57"
    value_range = f"Inputs!${cols[spec.value_label]}$8:${cols[spec.value_label]}$57"
    cost_range = f"Inputs!${cols[spec.cost_label]}$8:${cols[spec.cost_label]}$57"
    status_range = f"Inputs!${cols['Status']}$8:${cols['Status']}$57"
    risk_range = f"Inputs!${cols['Risk Flag']}$8:${cols['Risk Flag']}$57"
    dim_range = f"Inputs!${cols[spec.dimension_label]}$8:${cols[spec.dimension_label]}$57"
    for idx, month in enumerate(MONTH_VALUES, start=8):
        ws.cell(idx, 1, month)
        ws.cell(idx, 2, f'=SUMIFS({plan_range},{month_range},$A{idx})')
        ws.cell(idx, 3, f'=SUMIFS({actual_range},{month_range},$A{idx})')
        ws.cell(idx, 4, f'=SUMIFS({value_range},{month_range},$A{idx})')
        ws.cell(idx, 5, f'=SUMIFS({cost_range},{month_range},$A{idx})')
        ws.cell(idx, 6, f'=D{idx}*Assumptions!$B$10-E{idx}')
        ws.cell(idx, 7, f'=C{idx}-B{idx}')
        ws.cell(idx, 8, f'=IFERROR(G{idx}/ABS(B{idx}),0)')
        ws.cell(idx, 9, f'=IFERROR(F{idx}/E{idx},0)')
        ws.cell(idx, 10, f'=IFERROR((COUNTIFS({month_range},$A{idx},{status_range},"Complete")+COUNTIFS({month_range},$A{idx},{status_range},"Live")+COUNTIFS({month_range},$A{idx},{status_range},"Ready"))/COUNTIFS({month_range},$A{idx}),0)')
        ws.cell(idx, 11, f'=COUNTIFS({month_range},$A{idx},{risk_range},"REVIEW")')
        ws.cell(idx, 12, f'=IFERROR(XLOOKUP(MAXIFS({value_range},{month_range},$A{idx}),{value_range},{dim_range},""),"")')
        ws.cell(idx, 13, f'=IF(K{idx}>0,"Management review",IF(I{idx}>=Assumptions!$B$12,"Scale / continue","Improve economics"))')
    style_range(ws, 8, 1, 19, 13, FORMULA_FILL, alignment=Alignment(wrap_text=True))
    for row_idx in range(8, 20):
        for col_idx in range(2, 7):
            ws.cell(row_idx, col_idx).number_format = '"AED" #,##0;[Red]-"AED" #,##0;-'
        ws.cell(row_idx, 8).number_format = "0.0%;[Red](0.0%);-"
        ws.cell(row_idx, 9).number_format = "0.0%;[Red](0.0%);-"
        ws.cell(row_idx, 10).number_format = "0.0%"
    add_table(ws, f"tbl_{spec.slug}_calc", 7, 1, 19, 13)
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.conditional_formatting.add("K8:K19", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.conditional_formatting.add("I8:I19", CellIsRule(operator="lessThan", formula=["Assumptions!$B$12"], fill=PatternFill("solid", fgColor="FEE2E2")))


def create_dashboard(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    set_title(ws, spec.title, spec.management_question, 13)
    set_widths(ws, [16, 16, 3, 16, 16, 3, 16, 16, 3, 16, 16, 3, 18])
    cards = [
        ("Model Status", "=Checks!E18", "status"),
        ("Total Plan", "=SUM(Calc!B8:B19)", "money"),
        ("Total Actual", "=SUM(Calc!C8:C19)", "money"),
        ("Contribution ROI", '=LET(value,SUM(Calc!D8:D19),cost,SUM(Calc!E8:E19),IFERROR((value*Assumptions!$B$10-cost)/cost,0))', "percent"),
        ("Open Risks", "=SUM(Calc!K8:K19)", "count"),
        ("Best Month", '=IFERROR(INDEX(Calc!A8:A19,MATCH(MAX(Calc!I8:I19),Calc!I8:I19,0)),"")', "text"),
    ]
    positions = [(6, 1), (6, 4), (6, 7), (6, 10), (10, 1), (10, 4)]
    for (label, formula, kind), (r, c) in zip(cards, positions):
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + 1)
        ws.cell(r, c, label)
        ws.cell(r + 1, c, formula)
        style_range(ws, r, c, r + 2, c + 1, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True, vertical="center", horizontal="center"))
        ws.cell(r, c).font = Font(name=FONTS["mono"], size=8, bold=True, color=COLORS["muted"])
        ws.cell(r + 1, c).font = Font(name=FONTS["heading"], size=18, bold=True, color=COLORS["charcoal"])
        if kind == "money":
            ws.cell(r + 1, c).number_format = '"AED" #,##0;[Red]-"AED" #,##0;-'
        elif kind == "percent":
            ws.cell(r + 1, c).number_format = "0.0%"

    add_section(ws, 15, "Monthly performance summary", 13)
    monthly_headers = ["Month", "Plan", "Actual", "Revenue / Value", "Cost", spec.efficiency_label, "Completion %", "Open Risks", "Signal"]
    for c, header in enumerate(monthly_headers, start=1):
        ws.cell(16, c, header)
    for idx in range(17, 29):
        source = idx - 9
        formulas = [
            f"=Calc!A{source}",
            f"=Calc!B{source}",
            f"=Calc!C{source}",
            f"=Calc!D{source}",
            f"=Calc!E{source}",
            f"=Calc!I{source}",
            f"=Calc!J{source}",
            f"=Calc!K{source}",
            f"=Calc!M{source}",
        ]
        for c, formula in enumerate(formulas, start=1):
            ws.cell(idx, c, formula)
    style_range(ws, 16, 1, 16, 9, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 17, 1, 28, 9, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    for row_idx in range(17, 29):
        for col_idx in range(2, 5):
            ws.cell(row_idx, col_idx).number_format = '"AED" #,##0;[Red]-"AED" #,##0;-'
        ws.cell(row_idx, 6).number_format = "0.0%"
        ws.cell(row_idx, 7).number_format = "0.0%"

    chart = BarChart()
    chart.title = spec.chart_title
    chart.y_axis.title = "AED / score"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, max_col=3, min_row=16, max_row=28)
    cats = Reference(ws, min_col=1, min_row=17, max_row=28)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "J15")

    add_section(ws, 31, "Management read-out", 13)
    readout = [
        ("Performance", "=IF($B$7=\"OK\",\"The workbook checks are clear. Review the highest variance and strongest month before the next decision.\",\"Checks require review before this workbook is shared externally.\")"),
        ("Opportunity", "=IF($J$7>=Assumptions!$B$12,\"Current economics support scaling the strongest activity.\",\"Improve economics before adding budget or rollout pressure.\")"),
        ("Risk", "=IF($B$11>0,\"Open risk flags exist. Review Checks and Action_Plan.\",\"No open risk flags in the current sample data.\")"),
    ]
    for idx, (label, formula) in enumerate(readout, start=32):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, formula)
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=13)
    style_range(ws, 32, 1, 34, 13, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    ws.protection.sheet = True


def create_checks(wb: Workbook, spec: WorkbookSpec, cols: dict[str, str]) -> None:
    ws = wb.create_sheet("Checks")
    set_title(ws, "Audit / Checks", "Review these checks before using the workbook in a management, investor, franchise, or audit discussion.", 7)
    set_widths(ws, [34, 18, 18, 18, 14, 40, 22])
    headers = ["Review test", "Actual", "Expected", "Difference", "Status", "Fix hint", "Owner"]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    checks = [
        ("Missing record IDs", f'=COUNTIFS(Inputs!${cols[spec.item_label]}$8:${cols[spec.item_label]}$57,"<>",Inputs!$A$8:$A$57,"")', "0", "=B8-C8", '=IF(D8=0,"OK","REVIEW")', "Add a unique Record ID to every populated input row.", "Owner"),
        ("Missing owners", f'=COUNTIFS(Inputs!${cols[spec.item_label]}$8:${cols[spec.item_label]}$57,"<>",Inputs!${cols["Owner"]}$8:${cols["Owner"]}$57,"")', "0", "=B9-C9", '=IF(D9=0,"OK","REVIEW")', "Assign an owner before sharing the file.", "Owner"),
        ("Missing status", f'=COUNTIFS(Inputs!${cols[spec.item_label]}$8:${cols[spec.item_label]}$57,"<>",Inputs!${cols["Status"]}$8:${cols["Status"]}$57,"")', "0", "=B10-C10", '=IF(D10=0,"OK","REVIEW")', "Select a status from the dropdown.", "Owner"),
        ("Duplicate record IDs", '=SUMPRODUCT(--(Inputs!$A$8:$A$57<>""),--(COUNTIF(Inputs!$A$8:$A$57,Inputs!$A$8:$A$57)>1))', "0", "=B11-C11", '=IF(D11=0,"OK","REVIEW")', "Each row needs a unique ID for audit traceability.", "Finance Lead"),
        ("Open risk flags", f'=COUNTIF(Inputs!${cols["Risk Flag"]}$8:${cols["Risk Flag"]}$57,"REVIEW")', "0", "=B12-C12", '=IF(D12=0,"OK","REVIEW")', "Review risk rows in Inputs and Action_Plan.", "Management"),
        ("Material monthly variance", '=COUNTIF(Calc!H8:H19,">"&Assumptions!$B$11)', "0", "=B13-C13", '=IF(D13=0,"OK","REVIEW")', "Add commentary and corrective action for large variances.", "Finance Lead"),
        ("Low ROI / health months", '=COUNTIF(Calc!I8:I19,"<"&Assumptions!$B$12)', "0", "=B14-C14", '=IF(D14=0,"OK","REVIEW")', "Improve economics or pause low-return activity.", "Finance Lead"),
        ("Negative revenue / value", f'=COUNTIF(Inputs!${cols[spec.value_label]}$8:${cols[spec.value_label]}$57,"<0")', "0", "=B15-C15", '=IF(D15=0,"OK","REVIEW")', "Check signs and source reports.", "Finance Lead"),
        ("Missing cost on populated rows", f'=COUNTIFS(Inputs!${cols[spec.item_label]}$8:${cols[spec.item_label]}$57,"<>",Inputs!${cols[spec.cost_label]}$8:${cols[spec.cost_label]}$57,"")', "0", "=B16-C16", '=IF(D16=0,"OK","REVIEW")', "Add cost or use zero only where genuinely not applicable.", "Finance Lead"),
    ]
    for r, values in enumerate(checks, start=8):
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)
    ws.cell(18, 1, "Overall model status")
    ws.cell(18, 5, '=IF(COUNTIF(E8:E16,"REVIEW")=0,"OK","REVIEW")')
    ws.cell(18, 6, "If REVIEW appears, resolve failed checks before external use.")
    style_range(ws, 7, 1, 7, 7, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 8, 1, 18, 7, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    ws.conditional_formatting.add("E8:E18", FormulaRule(formula=['E8="REVIEW"'], fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.conditional_formatting.add("E8:E18", FormulaRule(formula=['E8="OK"'], fill=PatternFill("solid", fgColor="DCFCE7")))
    add_table(ws, f"tbl_{spec.slug}_checks", 7, 1, 18, 7)
    ws.protection.sheet = True


def create_definitions(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Definitions")
    set_title(ws, "Definitions / KPI Glossary", "KPI definitions, formula references, and source library.", 8)
    set_widths(ws, [28, 54, 36, 28, 38, 44, 18, 18])
    add_section(ws, 6, "Workbook KPI glossary", 8)
    headers = ["Term", "Definition", "Formula / Method", "Management use"]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    common = [
        ("Variance %", "Actual performance compared with plan or target.", "=(Actual - Plan) / ABS(Plan)", "Shows whether performance is materially off plan."),
        ("Risk Flag", "A row-level warning triggered by blocked status, critical priority, low ROI, or material variance.", "IF rule in Inputs", "Focuses management attention."),
        ("Model Status", "Overall status from the Checks sheet.", "No failed checks = OK", "Prevents sharing unchecked files."),
    ]
    definitions = common + spec.definitions
    for r, values in enumerate(definitions, start=8):
        for c, value in enumerate((*values, "Decision support"), start=1):
            ws.cell(r, c, value)
    style_range(ws, 7, 1, 7, 4, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 8, 1, 8 + len(definitions) - 1, 4, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    add_table(ws, f"tbl_{spec.slug}_definitions", 7, 1, 8 + len(definitions) - 1, 4)

    source_start = 12 + len(definitions)
    add_section(ws, source_start, "Reference sources used to design the framework", 8)
    source_headers = ["Area", "Source", "URL", "Applied in workbook"]
    for c, header in enumerate(source_headers, start=1):
        ws.cell(source_start + 1, c, header)
    for r, source in enumerate(SOURCE_LIBRARY, start=source_start + 2):
        for c, value in enumerate(source, start=1):
            ws.cell(r, c, value)
    style_range(ws, source_start + 1, 1, source_start + 1, 4, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, source_start + 2, 1, source_start + 1 + len(SOURCE_LIBRARY), 4, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    add_table(ws, f"tbl_{spec.slug}_sources", source_start + 1, 1, source_start + 1 + len(SOURCE_LIBRARY), 4)
    ws.protection.sheet = True


def create_scenarios(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Scenarios")
    set_title(ws, "Scenarios", "Base, Conservative, and Aggressive cases for quick management sensitivity review.", 9)
    set_widths(ws, [22, 18, 18, 18, 18, 18, 18, 22, 26])
    add_section(ws, 6, "Scenario controls", 9)
    headers = ["Scenario", "Spend / Cost Multiplier", "Revenue / Value Multiplier", "Margin / Conversion Uplift", "Scenario Cost", "Scenario Value", "Contribution Profit", "Scenario ROI", "Management Read"]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    scenarios = [
        ("Base Case", 1.00, 1.00, 0.00),
        ("Conservative Case", 0.90, 0.80, -0.05),
        ("Aggressive Case", 1.15, 1.25, 0.05),
    ]
    for r, (name, spend_mult, revenue_mult, uplift) in enumerate(scenarios, start=8):
        ws.cell(r, 1, name)
        ws.cell(r, 2, spend_mult)
        ws.cell(r, 3, revenue_mult)
        ws.cell(r, 4, uplift)
        ws.cell(r, 5, f"=SUM(Calc!E8:E19)*B{r}")
        ws.cell(r, 6, f"=SUM(Calc!D8:D19)*C{r}")
        ws.cell(r, 7, f"=F{r}*(Assumptions!$B$10+D{r})-E{r}")
        ws.cell(r, 8, f"=IFERROR(G{r}/E{r},0)")
        ws.cell(r, 9, f'=IF(H{r}>=Assumptions!$B$12,"Meets hurdle","Needs review")')
    style_range(ws, 7, 1, 7, 9, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 8, 1, 10, 9, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    for r in range(8, 11):
        for c in [2, 3, 4, 8]:
            ws.cell(r, c).number_format = "0.0%"
        for c in [5, 6, 7]:
            ws.cell(r, c).number_format = '"AED" #,##0;[Red]-"AED" #,##0;-'
        for c in [2, 3, 4]:
            ws.cell(r, c).fill = INPUT_FILL
            ws.cell(r, c).font = Font(name=FONTS["base"], color="0000FF")
    add_table(ws, f"tbl_{spec.slug}_scenarios", 7, 1, 10, 9)


def create_action_plan(wb: Workbook, spec: WorkbookSpec) -> None:
    ws = wb.create_sheet("Action_Plan")
    set_title(ws, "Action Plan / Recommendations", "Turn workbook signals into owned management action.", 10)
    set_widths(ws, [28, 48, 18, 14, 14, 15, 16, 16, 16, 30])
    headers = ["Issue / Opportunity", "Recommendation", "Owner", "Priority", "Due Date", "Status", "Impact AED", "Effort", "Days Overdue", "Evidence / Next Review"]
    for c, header in enumerate(headers, start=1):
        ws.cell(7, c, header)
    today = date(2026, 5, 13)
    for idx, action in enumerate(spec.actions, start=8):
        issue, recommendation, owner, priority = action
        ws.cell(idx, 1, issue)
        ws.cell(idx, 2, recommendation)
        ws.cell(idx, 3, owner)
        ws.cell(idx, 4, priority)
        ws.cell(idx, 5, today + timedelta(days=7 * (idx - 7)))
        ws.cell(idx, 6, "In Progress")
        ws.cell(idx, 7, 15000 * (idx - 7))
        ws.cell(idx, 8, "Medium")
        ws.cell(idx, 9, f'=IF(OR(E{idx}="",F{idx}="Complete"),"",MAX(0,TODAY()-E{idx}))')
        ws.cell(idx, 10, "Attach source report, screenshots, or manager sign-off.")
    for idx in range(8 + len(spec.actions), 18):
        ws.cell(idx, 9, f'=IF(OR(E{idx}="",F{idx}="Complete"),"",MAX(0,TODAY()-E{idx}))')
    style_range(ws, 7, 1, 7, 10, HEADER_FILL, Font(name=FONTS["base"], bold=True, color=COLORS["white"]))
    style_range(ws, 8, 1, 17, 10, PatternFill("solid", fgColor=COLORS["panel"]), alignment=Alignment(wrap_text=True))
    for r in range(8, 18):
        ws.cell(r, 5).number_format = "dd-mmm-yyyy"
        ws.cell(r, 7).number_format = '"AED" #,##0;[Red]-"AED" #,##0;-'
        for c in [1, 2, 3, 4, 5, 6, 7, 8, 10]:
            ws.cell(r, c).fill = INPUT_FILL
            ws.cell(r, c).font = Font(name=FONTS["base"], color="0000FF")
    for header, values in {"Owner": OWNER_VALUES, "Priority": PRIORITY_VALUES, "Status": STATUS_VALUES, "Effort": ["Low", "Medium", "High"]}.items():
        col = headers.index(header) + 1
        dv = list_validation(values)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}8:{get_column_letter(col)}17")
    ws.conditional_formatting.add("I8:I17", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FEE2E2")))
    add_table(ws, f"tbl_{spec.slug}_actions", 7, 1, 17, 10)


def finalize_workbook(wb: Workbook) -> None:
    # Remove the default empty worksheet if present.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.freeze_panes = ws.freeze_panes or "A6"
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    alignment = copy(cell.alignment)
                    alignment.vertical = alignment.vertical or "top"
                    cell.alignment = alignment
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass


def build_workbook(spec: WorkbookSpec) -> None:
    wb = Workbook()
    wb.properties.creator = "Ashmo.io"
    wb.properties.title = spec.title
    wb.properties.subject = spec.subtitle
    create_dashboard(wb, spec)
    create_readme(wb, spec)
    create_document_control(wb, spec)
    create_assumptions(wb, spec)
    _, cols = create_inputs(wb, spec)
    create_calc(wb, spec, cols)
    create_checks(wb, spec, cols)
    create_definitions(wb, spec)
    create_scenarios(wb, spec)
    create_action_plan(wb, spec)
    finalize_workbook(wb)
    spec.output_dir.mkdir(parents=True, exist_ok=True)
    wb.save(spec.path)


def rebuild_zip(specs: list[WorkbookSpec]) -> None:
    growth_specs = [spec for spec in specs if spec.folder == "growth"]
    zip_path = GROWTH_DIR / "restaurant-growth-toolkit-pack.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for spec in growth_specs:
            archive.write(spec.path, arcname=spec.filename)


ERROR_TOKENS = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"]
FUNCTIONS = ["SUMIFS", "COUNTIFS", "XLOOKUP", "INDEX", "MATCH", "IFERROR", "LET", "SUMPRODUCT", "MAXIFS"]
REQUIRED_SHEETS = ["Dashboard", "README", "Document_Control", "Assumptions", "Inputs", "Calc", "Checks", "Definitions", "Scenarios", "Action_Plan"]


def qa_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    formulas = 0
    functions = {name: 0 for name in FUNCTIONS}
    errors: list[str] = []
    tables = 0
    validations = 0
    cond_formats = 0
    charts = 0
    protected = 0
    for ws in wb.worksheets:
        tables += len(ws.tables)
        validations += len(ws.data_validations.dataValidation)
        cond_formats += len(ws.conditional_formatting)
        charts += len(getattr(ws, "_charts", []))
        protected += 1 if ws.protection.sheet else 0
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                value = cell.value
                if isinstance(value, str):
                    upper = value.upper()
                    if value.startswith("="):
                        formulas += 1
                        for func in FUNCTIONS:
                            if re.search(r"\b" + re.escape(func) + r"\s*\(", upper):
                                functions[func] += 1
                    for token in ERROR_TOKENS:
                        if token in upper:
                            errors.append(f"{ws.title}!{cell.coordinate}:{token}")
    return {
        "file": path.name,
        "path": str(path.relative_to(ROOT)),
        "size_kb": round(path.stat().st_size / 1024, 1),
        "sheets": wb.sheetnames,
        "required_sheets_present": all(sheet in wb.sheetnames for sheet in REQUIRED_SHEETS),
        "formulas": formulas,
        "formula_functions": {key: value for key, value in functions.items() if value},
        "tables": tables,
        "validations": validations,
        "conditional_formats": cond_formats,
        "charts": charts,
        "protected_sheets": protected,
        "formula_error_tokens": errors,
    }


def run_qa(specs: list[WorkbookSpec]) -> dict[str, Any]:
    reports = [qa_workbook(spec.path) for spec in specs]
    zip_path = GROWTH_DIR / "restaurant-growth-toolkit-pack.zip"
    zip_entries: list[str] = []
    if zip_path.exists():
        with zipfile.ZipFile(zip_path) as archive:
            zip_entries = sorted(archive.namelist())
    summary = {
        "workbooks_checked": len(reports),
        "all_required_sheets_present": all(report["required_sheets_present"] for report in reports),
        "total_formulas": sum(report["formulas"] for report in reports),
        "total_tables": sum(report["tables"] for report in reports),
        "total_validations": sum(report["validations"] for report in reports),
        "total_conditional_formats": sum(report["conditional_formats"] for report in reports),
        "total_charts": sum(report["charts"] for report in reports),
        "workbooks_with_errors": [report["file"] for report in reports if report["formula_error_tokens"]],
        "growth_zip_entries": zip_entries,
    }
    qa = {"generated_at": datetime.now().isoformat(timespec="seconds"), "summary": summary, "workbooks": reports}
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    QA_JSON.write_text(json.dumps(qa, indent=2), encoding="utf-8")
    lines = [
        "# Premium Workbook QA Report",
        "",
        f"Generated: {qa['generated_at']}",
        "",
        "## Summary",
        "",
        f"- Workbooks checked: {summary['workbooks_checked']}",
        f"- Required premium sheets present in every workbook: {summary['all_required_sheets_present']}",
        f"- Total formula cells: {summary['total_formulas']}",
        f"- Total Excel tables: {summary['total_tables']}",
        f"- Total data validations: {summary['total_validations']}",
        f"- Total conditional format collections: {summary['total_conditional_formats']}",
        f"- Total charts: {summary['total_charts']}",
        f"- Workbooks with formula error tokens: {', '.join(summary['workbooks_with_errors']) if summary['workbooks_with_errors'] else 'None'}",
        f"- Growth ZIP entries: {len(summary['growth_zip_entries'])}",
        "",
        "## Workbook Detail",
        "",
        "| Workbook | Sheets | Formulas | Tables | Validations | Conditional formats | Charts | Errors |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for report in reports:
        lines.append(
            f"| {report['file']} | {len(report['sheets'])} | {report['formulas']} | {report['tables']} | {report['validations']} | {report['conditional_formats']} | {report['charts']} | {len(report['formula_error_tokens'])} |"
        )
    QA_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return qa


def write_summary_report(specs: list[WorkbookSpec], qa: dict[str, Any]) -> None:
    lines = [
        "# Premium Workbook Rebuild Summary",
        "",
        f"Generated: {qa['generated_at']}",
        "",
        "## Portfolio Upgrade",
        "",
        "All 15 requested workbooks were rebuilt around a consistent management-model architecture:",
        "",
        "- `Dashboard` for board-level KPI cards, monthly summary, management read-out, and chart.",
        "- `README` for plain-language use instructions and workbook map.",
        "- `Document_Control` for version, owner, review status, and change log.",
        "- `Assumptions` for editable thresholds, margin, ROI, readiness, and workbook-specific drivers.",
        "- `Inputs` for structured editable data with Excel tables, filters, dropdowns, sample rows, and row-level formulas.",
        "- `Calc` for monthly SUMIFS/COUNTIFS/XLOOKUP-driven analysis.",
        "- `Checks` for missing inputs, duplicate IDs, open risks, material variance, low ROI, and negative values.",
        "- `Definitions` for KPI glossary and source library.",
        "- `Scenarios` for Base, Conservative, and Aggressive cases.",
        "- `Action_Plan` for recommendations, owners, priorities, due dates, and overdue logic.",
        "",
        "## Workbook-Specific Changes",
        "",
    ]
    for spec in specs:
        lines.extend(
            [
                f"### {spec.filename}",
                f"- Rebuilt as: {spec.title}",
                f"- Management question: {spec.management_question}",
                f"- Added domain inputs: {', '.join(spec.extra_headers)}",
                f"- Added decision focus: {spec.category}, {spec.efficiency_label}, checks, scenarios, and owner-led action plan.",
                "",
            ]
        )
    lines.extend(
        [
            "## QA Snapshot",
            "",
            f"- Workbooks checked: {qa['summary']['workbooks_checked']}",
            f"- Required sheet architecture present: {qa['summary']['all_required_sheets_present']}",
            f"- Formula cells: {qa['summary']['total_formulas']}",
            f"- Excel tables: {qa['summary']['total_tables']}",
            f"- Data validations: {qa['summary']['total_validations']}",
            f"- Conditional formatting collections: {qa['summary']['total_conditional_formats']}",
            f"- Charts: {qa['summary']['total_charts']}",
            f"- Formula error tokens found: {', '.join(qa['summary']['workbooks_with_errors']) if qa['summary']['workbooks_with_errors'] else 'None'}",
            "",
            "## Recommended Future Improvements",
            "",
            "- Build rendered preview images for each workbook dashboard and show them on the website.",
            "- Add downloadable sample PDF board packs generated from the Dashboard sheets.",
            "- Add a paid/pro version with branch-level consolidation and actual POS/CRM import mappings.",
            "- Add workbook-specific video walkthroughs for non-finance users.",
            "- Add localized currency/tax variants for UAE, Saudi Arabia, Qatar, and India operators.",
            "- Add a versioned changelog page so template users know what improved between releases.",
        ]
    )
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_all() -> tuple[list[WorkbookSpec], dict[str, Any]]:
    specs = add_missing_specs(extend_specs(make_specs()))
    for spec in specs:
        build_workbook(spec)
    rebuild_zip(specs)
    qa = run_qa(specs)
    write_summary_report(specs, qa)
    return specs, qa


def main() -> None:
    parser = argparse.ArgumentParser(description="Build or QA premium Ashmo workbooks.")
    parser.add_argument("--qa-only", action="store_true", help="Run QA without rebuilding.")
    args = parser.parse_args()
    specs = add_missing_specs(extend_specs(make_specs()))
    if args.qa_only:
        qa = run_qa(specs)
        write_summary_report(specs, qa)
    else:
        _, qa = build_all()
    print(json.dumps(qa["summary"], indent=2))


if __name__ == "__main__":
    main()
